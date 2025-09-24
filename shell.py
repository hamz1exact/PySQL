import sys
import os
import time
import json
import csv
import importlib
import atexit
import re
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import argparse


# Enhanced terminal UI imports
try:
    from prompt_toolkit import prompt
    from prompt_toolkit.completion import WordCompleter, Completer, Completion
    from prompt_toolkit.history import FileHistory
    from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
    from prompt_toolkit.shortcuts import print_formatted_text
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.styles import Style
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.keys import Keys
    PROMPT_TOOLKIT_AVAILABLE = True
except ImportError:
    PROMPT_TOOLKIT_AVAILABLE = False
    print("Warning: prompt_toolkit not available. Install with: pip install prompt-toolkit")

# Module registry for hot-reloading
MODULE_REGISTRY = {}

def register_module(name, module):
    """Register a module for hot-reloading"""
    MODULE_REGISTRY[name] = module

def reload_modules():
    """Hot-reload all registered modules"""
    reloaded = []
    failed = []
    
    for name, module in MODULE_REGISTRY.items():
        try:
            importlib.reload(module)
            reloaded.append(name)
        except Exception as e:
            failed.append((name, str(e)))
    
    return reloaded, failed

# Try to import your existing modules
try:
    import database_manager
    import engine
    import executor
    import datatypes
    
    # Register modules for hot-reloading
    register_module('database_manager', database_manager)
    register_module('engine', engine)
    register_module('executor', executor)
    register_module('datatypes', datatypes)
    
    from engine import db_manager, Lexer, Parser
    from executor import execute
    
except ImportError as e:
    print(f"Warning: Could not import SQL engine modules: {e}")
    print("Make sure your engine.py, executor.py, and database_manager.py are in the Python path")
    
    # Create mock objects for testing
    class MockDBManager:
        def __init__(self):
            self.active_db = {}
            self.active_db_name = None
            self.databases = []
        
        def save_database_file(self):
            pass
    
    class MockLexer:
        def __init__(self, query):
            self.tokens = [("SELECT", "SELECT")]
    
    class MockParser:
        def __init__(self, tokens):
            self.tokens = tokens
        
        def parse_select_statement(self):
            return {"type": "SELECT", "mock": True}
    
    def mock_execute(ast, database):
        return [{"message": "Mock execution - engine modules not available"}]
    
    db_manager = MockDBManager()
    Lexer = MockLexer
    Parser = MockParser
    execute = mock_execute


class Config:
    """Enhanced configuration with hot-reload support"""
    def __init__(self):
        self.show_query_time = True
        self.show_row_count = True
        self.max_column_width = 50
        self.null_display = 'NULL'
        self.prompt_format = '{db}> '
        self.pager = 'less'
        self.history_size = 1000
        self.auto_commit = True
        self.echo_queries = False
        self.table_format = 'ascii'  # ascii, markdown, csv
        self.timing = True
        self.last_result = None
        self.last_columns = None
        self.auto_reload = True  # Enable auto-reload on file changes
        self.reload_on_error = False  # Reload modules on execution error
        self.auto_detect_wide = False   # Default to false - user must explicitly enable
        self.wide_table_threshold = 10  # columns threshold
        self.display_mode = 'auto'  # 'auto', 'force_normal', 'force_wide'
        
        # Load config from file if exists
        self.config_file = Path.home() / '.myshell_config.json'
        self.load_config()
    
    def load_config(self):
        """Load configuration from file"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config_data = json.load(f)
                    for key, value in config_data.items():
                        if hasattr(self, key):
                            setattr(self, key, value)
        except Exception as e:
            print(f"Warning: Could not load config file: {e}")
    
    def save_config(self):
        """Save configuration to file"""
        try:
            config_data = {
                'show_query_time': self.show_query_time,
                'show_row_count': self.show_row_count,
                'max_column_width': self.max_column_width,
                'null_display': self.null_display,
                'prompt_format': self.prompt_format,
                'pager': self.pager,
                'history_size': self.history_size,
                'auto_commit': self.auto_commit,
                'echo_queries': self.echo_queries,
                'table_format': self.table_format,
                'timing': self.timing,
                'auto_reload': self.auto_reload,
                'reload_on_error': self.reload_on_error,
                'auto_detect_wide': self.auto_detect_wide,
                'wide_table_threshold': self.wide_table_threshold,
                'display_mode': self.display_mode
            }
            with open(self.config_file, 'w') as f:
                json.dump(config_data, f, indent=2)
        except Exception as e:
            print(f"Warning: Could not save config file: {e}")


class FilteredFileHistory(FileHistory):
    """File history that filters out meta commands"""
    
    def __init__(self, filename):
        super().__init__(filename)
    
    def append_string(self, string):
        """Only append SQL queries, not meta commands"""
        if not string.strip().startswith('\\'):
            super().append_string(string)


class SQLCompleter(Completer):
    """Enhanced SQL completer with keywords, table names, and column names"""
    
    def __init__(self, db_manager=None):
        self.db_manager = db_manager
        
        # Comprehensive SQL keywords
        self.sql_keywords = [
            # DML
            'SELECT', 'FROM', 'WHERE', 'INSERT', 'INTO', 'VALUES', 'UPDATE', 'SET', 'DELETE',
            # DDL
            'CREATE', 'DROP', 'ALTER', 'TABLE', 'DATABASE', 'INDEX', 'VIEW', 'TRIGGER',
            # Constraints
            'PRIMARY', 'KEY', 'UNIQUE', 'NOT', 'NULL', 'DEFAULT',
            'CHECK', 'CONSTRAINT',
            # Data Types
            'INT', 'INTEGER', 'VARCHAR', 'CHAR', 'TEXT',
            'DATE', 'TIME', 'TIMESTAMP', 'BOOLEAN', 'FLOAT', 'DOUBLE',
            # Operators and Functions
            'AND', 'OR', 'IN', 'LIKE', 'BETWEEN', 'IS', 'EXISTS', 'CASE', 'WHEN', 'THEN',
            'ELSE', 'END', 'AS', 'DISTINCT',
            # Grouping and Ordering
            'ORDER', 'BY', 'GROUP', 'HAVING', 'ASC', 'DESC',
            # Aggregate Functions
            'COUNT', 'SUM', 'AVG', 'MIN', 'MAX',
            # String Functions
            'LENGTH', 'SUBSTR', 'SUBSTRING', 'UPPER', 'LOWER',
            
            'REPLACE', 'CONCAT',
            # Numeric Functions
            'ABS', 'ROUND', 'CEIL', 'CEILING', 'FLOOR',
            # Date Functions
            'NOW', 'CURRENT_DATE', 'CURRENT_TIME',
            
            'YEAR', 'MONTH', 'DAY', 'HOUR', 'MINUTE', 'SECOND',
            # Other Keywords
            'USE', 'SHOW', 'DESC', 'UNION', 'EXCEPT', 'INTERSECT',
            
            'WITH',  'LIMIT', 'OFFSET', 'FETCH',
            
            'GRANT', 'REVOKE', 'PRIVILEGES', 'TO', 'PUBLIC',
            
            'TRUE', 'FALSE', 'UNKNOWN', 'RETURNING', 'CONSTRANTS', "VIEWS"
        ]
    
    def get_completions(self, document, complete_event):
        word = document.get_word_before_cursor()
        
        # SQL Keywords
        for keyword in self.sql_keywords:
            if keyword.lower().startswith(word.lower()):
                yield Completion(keyword, start_position=-len(word))
        
        # Table names
        try:
            if self.db_manager and hasattr(self.db_manager, 'active_db') and self.db_manager.active_db:
                for table_name in self.db_manager.active_db.keys():
                    if table_name.lower().startswith(word.lower()):
                        yield Completion(table_name, start_position=-len(word))
                
                # Column names (from all tables)
                for table_name, table_obj in self.db_manager.active_db.items():
                    if hasattr(table_obj, 'schema'):
                        for column_name in table_obj.schema.keys():
                            if column_name.lower().startswith(word.lower()):
                                yield Completion(column_name, start_position=-len(word))
        except:
            pass


class EnhancedHistoryManager:
    """Enhanced history management with prompt_toolkit support"""
    def __init__(self, config: Config):
        self.config = config
        self.history_file = Path.home() / '.myshell_history'
        
        if PROMPT_TOOLKIT_AVAILABLE:
            self.file_history = FilteredFileHistory(str(self.history_file))
        else:
            self.setup_readline_history()
    
    def setup_readline_history(self):
        """Fallback to readline history"""
        try:
            import readline
            readline.set_history_length(self.config.history_size)
            
            if self.history_file.exists():
                readline.read_history_file(str(self.history_file))
            
            atexit.register(self.save_readline_history)
        except ImportError:
            print("Warning: readline not available, history disabled")
    
    def save_readline_history(self):
        """Save readline history"""
        try:
            import readline
            readline.write_history_file(str(self.history_file))
        except:
            pass


class TableFormatter:
    """Enhanced table formatting with better visual appeal"""
    
    def __init__(self, config: Config):
        self.config = config
    
    def format_table(self, rows: List[Dict], columns: List[str] = None) -> str:
        """Format table data for display"""
        if not rows:
            return "Empty result set"
        
        if columns is None:
            columns = list(rows[0].keys()) if rows else []
        
        if self.config.table_format == 'csv':
            return self._format_csv(rows, columns)
        elif self.config.table_format == 'markdown':
            return self._format_markdown(rows, columns)
        else:
            return self._format_ascii(rows, columns)
    
    def _format_ascii(self, rows: List[Dict], columns: List[str]) -> str:
        """Enhanced ASCII table with display mode support"""
        if not rows:
            return "Empty result set"
        
        # Check display mode
        num_columns = len(columns)
        
        if self.config.display_mode == 'force_normal':
            # Always show normal table
            return self._format_normal_table(rows, columns)
        elif self.config.display_mode == 'force_wide':
            # Always show wide format
            return self._format_wide_table(rows, columns)
        else:  # auto mode
            if self.config.auto_detect_wide and num_columns > self.config.wide_table_threshold:
                return self._format_wide_table(rows, columns)
            else:
                return self._format_normal_table(rows, columns)
    
    def _format_normal_table(self, rows: List[Dict], columns: List[str]) -> str:
        """Format as normal ASCII table"""
        col_widths = {}
        for col in columns:
            col_widths[col] = min(len(col), self.config.max_column_width)
            
        for row in rows:
            for col in columns:
                value_str = self._format_value(row.get(col))
                col_widths[col] = max(col_widths[col], min(len(value_str), self.config.max_column_width))
        
        lines = []
        
        # Top border
        border_parts = []
        for col in columns:
            border_parts.append('-' * (col_widths[col] + 2))
        lines.append('+' + '+'.join(border_parts) + '+')
        
        # Header
        header_parts = []
        for col in columns:
            header_text = col[:col_widths[col]].ljust(col_widths[col])
            header_parts.append(f' {header_text} ')
        lines.append('|' + '|'.join(header_parts) + '|')
        
        # Header separator
        lines.append('+' + '+'.join(border_parts) + '+')
        
        # Data rows
        for row in rows:
            row_parts = []
            for col in columns:
                value = row.get(col)
                value_str = self._format_value(value)
                
                if len(value_str) > col_widths[col]:
                    value_str = value_str[:col_widths[col]-3] + '...'
                
                value_str = value_str.ljust(col_widths[col])
                row_parts.append(f' {value_str} ')
            lines.append('|' + '|'.join(row_parts) + '|')
        
        # Bottom border
        lines.append('+' + '+'.join(border_parts) + '+')
        
        return '\n'.join(lines)
    
    def _format_wide_table(self, rows: List[Dict], columns: List[str]) -> str:
        """Format wide tables in vertical format"""
        lines = []
        
        for i, row in enumerate(rows, 1):
            lines.append(f"*************************** {i}. row ***************************")
            
            # Find the maximum column name length for alignment
            max_col_len = max(len(col) for col in columns) if columns else 0
            
            for col in columns:
                value = self._format_value(row.get(col))
                col_name = col.rjust(max_col_len)
                lines.append(f"{col_name}: {value}")
            
            if i < len(rows):  # Add separator between records
                lines.append("")
        
        return '\n'.join(lines)
    
    def _format_markdown(self, rows: List[Dict], columns: List[str]) -> str:
        """Enhanced Markdown table format"""
        if not rows:
            return "| Empty result set |"
        
        lines = []
        
        # Header
        header = '| ' + ' | '.join(f"**{col}**" for col in columns) + ' |'
        lines.append(header)
        
        # Separator
        separator = '| ' + ' | '.join([':---:'] * len(columns)) + ' |'
        lines.append(separator)
        
        # Data rows
        for row in rows:
            row_data = []
            for col in columns:
                value = self._format_value(row.get(col))
                # Escape pipe characters in markdown
                value = value.replace('|', '\\|')
                row_data.append(value)
            lines.append('| ' + ' | '.join(row_data) + ' |')
        
        return '\n'.join(lines)
    
    def _format_csv(self, rows: List[Dict], columns: List[str]) -> str:
        """CSV format implementation"""
        lines = []
        lines.append(','.join(columns))
        
        for row in rows:
            row_data = []
            for col in columns:
                value = self._format_value(row.get(col))
                if ',' in value or '"' in value or '\n' in value:
                    value = '"' + value.replace('"', '""') + '"'
                row_data.append(value)
            lines.append(','.join(row_data))
        
        return '\n'.join(lines)
    
    def _format_value(self, value) -> str:
        """Format a single value for display"""
        if value is None:
            return self.config.null_display
        elif isinstance(value, bool):
            return 'TRUE' if value else 'FALSE'
        elif isinstance(value, (int, float)):
            return str(value)
        else:
            return str(value)


class EnhancedSQLShell:
    """Enhanced SQL shell with hot-reloading and better terminal UI"""
    
    def __init__(self):
        self.config = Config()
        self.history = EnhancedHistoryManager(self.config)
        self.formatter = TableFormatter(self.config)
        self.query_count = 0
        self.start_time = datetime.now()
        
        # Setup completions
        if PROMPT_TOOLKIT_AVAILABLE:
            self.completer = SQLCompleter(db_manager)
            self.style = Style.from_dict({})  # No colors
            
            # Setup key bindings for Ctrl+Enter

        
        # Enhanced command mappings
        self.meta_commands = {
            '\\q': self._cmd_quit,
            '\\quit': self._cmd_quit,
            '\\exit': self._cmd_quit,
            '\\h': self._cmd_help,
            '\\help': self._cmd_help,
            '\\?': self._cmd_help,
            '\\l': self._cmd_list_databases,
            '\\list': self._cmd_list_databases,
            '\\d': self._cmd_describe_tables,
            '\\dt': self._cmd_describe_tables,
            '\\c': self._cmd_connect,
            '\\connect': self._cmd_connect,
            '\\use': self._cmd_connect,
            '\\timing': self._cmd_toggle_timing,
            '\\echo': self._cmd_toggle_echo,
            '\\set': self._cmd_set_config,
            '\\show': self._cmd_show_config,
            '\\history': self._cmd_show_history,
            '\\clear': self._cmd_clear,
            '\\cls': self._cmd_clear,
            '\\export': self._cmd_export,
            '\\import': self._cmd_import,
            '\\version': self._cmd_version,
            '\\status': self._cmd_status,
            '\\modules': self._cmd_list_modules,
            '\\debug': self._cmd_debug_mode,
            '\\wide' : self._cmd_wide,
            '\\cols': self._cmd_cols,
            '\\normal': self._cmd_set_normal,
            '\\force': self._cmd_set_normal,
            '\\auto': self._cmd_set_auto,
            '\\csv': self._cmd_show_csv_style,
            '\\columns': self._cmd_show_columns_only,
            '\\schema':  self._cmd_show_columns_only
        }
    
    def _get_connection_info(self) -> str:
        """Get current connection information"""
        try:
            current_db = getattr(db_manager, 'active_db_name', None)
            if current_db:
                filename = os.path.basename(current_db)
                if filename.endswith('.su'):
                    filename = filename[:-3]
                return f"Connected to: {filename}"
            else:
                return "No database selected. Use USE database_name to connect."
        except:
            return "Database manager not available"
    
    def get_prompt_text(self) -> str:
        """Generate the command prompt text"""
        try:
            db_name = getattr(db_manager, 'active_db_name', None) or 'myshell'
            
            if db_name and db_name != 'myshell':
                filename = os.path.basename(db_name)
                if filename.endswith('.su'):
                    filename = filename[:-3]
                db_name = filename
        except:
            db_name = 'myshell'
        
        return f"{db_name}> "
    
    def get_formatted_prompt(self):
        """Get formatted prompt for prompt_toolkit"""
        if PROMPT_TOOLKIT_AVAILABLE:
            db_name = 'myshell'
            try:
                current_db = getattr(db_manager, 'active_db_name', None)
                if current_db:
                    filename = os.path.basename(current_db)
                    if filename.endswith('.su'):
                        filename = filename[:-3]
                    db_name = filename
            except:
                pass
            
            return f"{db_name}> "
        else:
            return self.get_prompt_text()
    
    def _strip_comments(self, query: str) -> str:
        """Remove SQL comments from query"""
        lines = query.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # Find comment start
            comment_pos = line.find('--')
            if comment_pos != -1:
                # Keep everything before the comment
                line = line[:comment_pos].rstrip()
            
            if line.strip():  # Only add non-empty lines
                cleaned_lines.append(line)
        
        return '\n'.join(cleaned_lines)
    
    def _split_multiple_queries(self, query_text: str) -> List[str]:
        """Split multiple queries separated by semicolons"""
        # Remove comments first
        cleaned_query = self._strip_comments(query_text)
        
        # Split by semicolon and filter empty queries
        queries = []
        for q in cleaned_query.split(';'):
            q = q.strip()
            if q:
                queries.append(q + ';')  # Add semicolon back
        
        return queries
    
    def run(self):
        
        query_buffer = []
        
        while True:
            try:
                # Get input with enhanced prompting
                if PROMPT_TOOLKIT_AVAILABLE:
                    if query_buffer:
                        continuation_prompt = "    -> "
                        line = prompt(continuation_prompt, 
                                    style=self.style,
                                    history=self.history.file_history,
                                    auto_suggest=AutoSuggestFromHistory(),
            completer=self.completer).strip()
                                    
                    else:
                        line = prompt(self.get_formatted_prompt(),
                                    style=self.style,
                                    history=self.history.file_history,
                                    auto_suggest=AutoSuggestFromHistory(),
                                    completer=self.completer).strip()
                else:
                    # Fallback to basic input
                    if query_buffer:
                        prompt_text = "    -> "
                    else:
                        prompt_text = self.get_prompt_text()
                    line = input(prompt_text).strip()
                
                if not line:
                    continue
                
                # Handle meta-commands
                if line.startswith('\\'):
                    if query_buffer:
                        print("Discarding incomplete query")
                        query_buffer = []
                    
                    self._handle_meta_command(line)
                    continue
                
                # Build query buffer
                query_buffer.append(line)
                
                # Check if we have complete queries (ending with semicolon)
                full_query = ' '.join(query_buffer)
                if ';' in full_query:
                    # Execute all complete queries
                    queries = self._split_multiple_queries(full_query)
                    
                    for query in queries:
                        if query.strip():
                            self._execute_query(query)
                    
                    query_buffer = []
                
            except KeyboardInterrupt:
                print("^C - Query cancelled")
                query_buffer = []
                continue
            except EOFError:
                print("\nGoodbye!")
                break
            except Exception as e:
                print(f"Unexpected error: {e}")
    
    def _handle_meta_command(self, command: str):
        """Handle meta-commands with enhanced error reporting"""
        parts = command.split()
        cmd = parts[0].lower()
        args = parts[1:] if len(parts) > 1 else []
        
        if cmd in self.meta_commands:
            try:
                self.meta_commands[cmd](args)
            except Exception as e:
                print(f"Command failed: {e}")
                print(f"Details: {traceback.format_exc()}")
        else:
            print(f"Unknown command: {cmd}")
            print("Type \\h for help")
    
    def _execute_query(self, query: str):
        """Execute SQL query with enhanced error handling and hot-reload support"""
        query = query.strip()
        if not query:
            return
        
        # Strip comments
        query = self._strip_comments(query)
        if not query:
            return
        
        if self.config.echo_queries:
            print(query)
        
        start_time = time.time()
        
        try:
            # Get current database
            database = db_manager.active_db
            
            # Use your existing Lexer and Parser
            lexer = Lexer(query)
            parser = Parser(lexer.tokens)
            
            if not lexer.tokens:
                return
            
            token_type = lexer.tokens[0][0]
            next_token_type = lexer.tokens[1][0] if len(lexer.tokens) > 1 else None
                        
            # Parse and execute based on query type
            if token_type == "SELECT":
                from sql_ast import UnionExpression, IntersectExpression, ExceptExpression
                ast = parser.parse_select_statement()
                if isinstance(ast, (UnionExpression, IntersectExpression, ExceptExpression)):
                    result = ast.evaluate()
                else:
                    result = execute(ast, db_manager)
                self._handle_select_result(result, start_time)
                
            elif token_type == "INSERT":
                from sql_ast import SelectStatement, UpdateStatement
                ast = parser.parse_insert_statement()
                result = execute(ast, database)
                if ast.returned_cols:
                    returning_result = ast.returned_cols.evaluate(result, database)
                    self._handle_select_result(returning_result, start_time)
                db_manager.save_database_file()
                self._handle_modify_result("INSERT", start_time)
                
            elif token_type == "UPDATE":
                ast = parser.parse_update_statement()
                result = execute(ast, database)
                if ast.returned_columns:
                    returning_result = ast.returned_columns.evaluate(result, database)
                    self._handle_select_result(returning_result, start_time)
                db_manager.save_database_file()
                self._handle_modify_result("UPDATE", start_time)
                
            elif token_type == "DELETE":
                ast = parser.parse_delete_statement()
                result = execute(ast, database)
                if ast.returned_columns:
                    returning_result = ast.returned_columns.evaluate(result, database)
                    self._handle_select_result(returning_result, start_time)
                db_manager.save_database_file()
                self._handle_modify_result("DELETE", start_time)
            
            elif token_type == "WITH":
                ast = parser.parse_cte()
                result = ast.execute(db_manager)
                self._handle_select_result(result, start_time)
            
            elif token_type == "SHOW":
                from sql_ast import ShowConstraints
                ast = parser.parse_request_statement()
                if isinstance(ast, ShowConstraints):
                    result = ast.evaluate()
                    self._handle_select_result(result, start_time)
                else:
                    result = ast
                
            elif token_type == "CALL":
                ast = parser.parse_calling_expression()
                result = execute(ast, db_manager)
                self._handle_select_result(result.evaluate(), start_time)
              
            elif token_type == "REFRESH":
                ast = parser.parse_refresh_mv()
                result = execute(ast, db_manager)
                
            
            
            elif token_type == "TRUNCATE":
                ast = parser.parse_truncate_table()
                result = execute(ast, db_manager)
                self._handle_ddl_result("TRUNCATE TABLE", start_time)
                db_manager.save_database_file()
            elif token_type == "CREATE":
                if next_token_type == "DATABASE":
                    ast = parser.parse_create_database()
                    result = execute(ast, db_manager)
                elif next_token_type == "TABLE":
                    ast = parser.parse_create_table()
                    execute(ast, db_manager)
                    db_manager.save_database_file()
                    self._handle_ddl_result("CREATE TABLE", start_time)
                elif next_token_type == "HIGH_PRIORITY_OPERATOR" or next_token_type == "VIEW" or next_token_type == "MATERIALIZED":
                    ast = parser.create_view()
                    result = execute(ast, db_manager)
                    db_manager.save_database_file()
                    db_manager.update_cache()
                else:
                    
                    print("Unsupported CREATE statement")
                    
            elif token_type == "DROP":
                if next_token_type == "DATABASE":
                    ast = parser.parse_drop_database()
                    result = execute(ast, db_manager)
                    self._handle_ddl_result("DROP DATABASE", start_time)
                    db_manager.update_cache()
                    db_manager.save_database_file()
                elif next_token_type == "TABLE":
                    ast = parser.parse_drop_table()
                    result = execute(ast, db_manager)
                    db_manager.save_database_file()
                    self._handle_ddl_result("DROP TABLE", start_time)
                elif next_token_type == "VIEW":
                    ast = parser.parse_drop_view()
                    result = execute(ast, db_manager)
                    self._handle_ddl_result("DROP VIEW", start_time)
                elif next_token_type == "MATERIALIZED":
                    ast = parser.parse_drop_mtv()
                    result = execute(ast, db_manager)
                    self._handle_ddl_result("DROP MATERIALIZED VIEW", start_time)
                else:
                    print("Unsupported DROP statement")
                    
            elif token_type == "ALTER":
                ast = parser.parse_alter_table()
                result = ast.execute(db_manager)
                self._handle_ddl_result("ALTER TABLE", start_time)
                db_manager.save_database_file()
            elif token_type == "USE":
                ast = parser.parse_use_statement()
                result = execute(ast, db_manager)
                self._handle_use_result("USE", start_time)
                
            else:
                print(f"Unsupported statement type '{token_type}'")
            
            self.query_count += 1
            
        except Exception as e:
            execution_time = time.time() - start_time
            print(f"Query execution failed: {str(e)}")
            
            if self.config.timing:
                time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
                print(f"Time: {time_str}")
            
            # Auto-reload on error if enabled
            if self.config.reload_on_error:
                print("Auto-reloading modules...")

        
        print()  # Empty line for readability
    
    def _handle_select_result(self, result, start_time):
        """Handle SELECT query results"""
        execution_time = time.time() - start_time
        
        # Store the result
        self.config.last_result = result
        
        if result:  # Assuming result is a list of dictionaries for SELECT
            formatted_table = self.formatter.format_table(result)
            print(formatted_table)
            
            # Show summary
            if self.config.show_row_count:
                row_count = len(result)
                plural = 's' if row_count != 1 else ''
                print(f"\n({row_count} row{plural})")
        else:
            print("Empty result set")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"Time: {time_str}")
    
    def _handle_modify_result(self, operation, start_time):
        """Handle INSERT/UPDATE/DELETE results"""
        execution_time = time.time() - start_time
        print(f"{operation} operation completed")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"Time: {time_str}")
    
    def _handle_ddl_result(self, operation, start_time):
        """Handle CREATE/DROP/ALTER results"""
        execution_time = time.time() - start_time
        print(f"{operation} completed successfully")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"Time: {time_str}")
    
    def _handle_use_result(self, operation, start_time):
        """Handle USE database results"""
        execution_time = time.time() - start_time
        
        try:
            current_db_name = getattr(db_manager, 'active_db_name', 'unknown')
            if current_db_name and current_db_name != 'unknown':
                filename = os.path.basename(current_db_name)
                if filename.endswith('.su'):
                    filename = filename[:-3]
                current_db_name = filename
            
            print(f"Switched to database: {current_db_name}")
        except:
            print("Database switch completed")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"Time: {time_str}")
    
    # Enhanced Meta-command implementations
    def _cmd_quit(self, args):
        """Quit the shell with cleanup"""
        self.config.save_config()
        print(f"Session lasted {datetime.now() - self.start_time}")
        sys.exit(0)
    
    def _cmd_help(self, args):
        """Show comprehensive help information"""
        help_text = f"""

Meta Commands:
  \\h, \\help              Show this help
  \\q, \\quit, \\exit      Quit the shell
  \\l, \\list              List all databases
  \\d, \\dt                List tables in current database
  \\c <db>, \\connect <db>  Connect to database
  \\clear, \\cls           Clear screen
  

Display Mode Commands:
  \\normal, \\force       Set normal table display (persistent)
  \\auto                  Set auto table display mode (persistent)
  \\wide vertical         Show last result in vertical format
  \\cols                  Show specific columns from last result

Configuration:
  \\timing                Toggle query timing on/off
  \\echo                  Toggle query echo on/off
  \\set <option> <value>  Set configuration option
  \\show [option]         Show configuration

Data Operations:
  \\export <format> <file> Export last SELECT result
  \\import <file>         Import and execute SQL file
  \\history               Show command history

Information:
  \\version               Show version information
  \\status                Show session status

Configuration Options:
  table_format           ascii, markdown, csv
  max_column_width       Maximum width for table columns
  null_display           How to display NULL values
  auto_detect_wide       Enable/disable wide table detection (true/false)
  wide_table_threshold   Column count threshold for wide tables
  display_mode          'auto', 'force_normal', 'force_wide'
  auto_reload            Auto-reload modules on file changes
  reload_on_error        Auto-reload on execution errors

Multi-Query Support:
  • Separate multiple queries with semicolons (;)
  • Press Enter to execute complete queries
  • Press Ctrl+Enter to add new line within query
  • Comments with -- are supported

Tips:
  • End queries with semicolon (;)
  • Multi-line and multi-query support
  • Ctrl+C cancels current query
  • Use arrow keys for history navigation
  • Tab completion for SQL keywords, tables, and columns
"""
        print(help_text)
    
    def _cmd_list_databases(self, args):
        """List all databases with enhanced display"""
        print("Available Databases:\n")
        
        try:
            if not hasattr(db_manager, 'databases') or not db_manager.databases:
                print("No databases found")
                return
            
            databases_info = []
            for db_file in db_manager.databases:
                try:
                    filename = os.path.basename(db_file)
                    if filename.endswith('.su'):
                        filename = filename[:-3]
                    
                    # Determine status
                    if db_file == getattr(db_manager, 'active_db_name', None):
                        status = "Online"
                    else:
                        status = "Available"
                    
                    # Get file size
                    try:
                        size = os.path.getsize(db_file)
                        size_str = f"{size:,} bytes" if size < 1024 else f"{size//1024:,} KB"
                    except:
                        size_str = "Unknown"
                    
                    databases_info.append({
                        'Database': filename,
                        'Status': status,
                        'Size': size_str,
                        'Path': db_file
                    })
                
                except Exception as e:
                    databases_info.append({
                        'Database': str(db_file),
                        'Status': "ERROR",
                        'Size': "N/A",
                        'Path': str(e)
                    })
            
            if databases_info:
                formatted_table = self.formatter.format_table(databases_info, ['Database', 'Status', 'Size'])
                print(formatted_table)
            else:
                print("No databases available")
        
        except Exception as e:
            print(f"Failed to list databases: {e}")
    
    def _cmd_describe_tables(self, args):
        """List tables in current database with enhanced info"""
        try:
            current_db = getattr(db_manager, 'active_db', None)
            if not current_db:
                print("No database selected")
                print("Use \\c <database_name> to connect to a database")
                return
            
            print("Tables in current database:\n")
            
            tables_info = []
            for table_name, table_obj in current_db.items():
                try:
                    row_count = len(table_obj.rows) if hasattr(table_obj, 'rows') else 0
                    col_count = len(table_obj.schema) if hasattr(table_obj, 'schema') else 0
                    
                    # Get column info
                    columns = []
                    if hasattr(table_obj, 'schema'):
                        for col_name, col_type in table_obj.schema.items():
                            columns.append(f"{col_name}({col_type})")
                    
                    tables_info.append({
                        'Table': table_name,
                        'Rows': f"{row_count:,}",
                        'Columns': str(col_count),
                        'Schema': ', '.join(columns[:3]) + ('...' if len(columns) > 3 else '')
                    })
                
                except Exception as e:
                    tables_info.append({
                        'Table': table_name,
                        'Rows': 'Error',
                        'Columns': 'Error',
                        'Schema': str(e)[:50]
                    })
            
            if tables_info:
                formatted_table = self.formatter.format_table(tables_info, ['Table', 'Rows', 'Columns', 'Schema'])
                print(formatted_table)
            else:
                print("No tables found in current database")

        except Exception as e:
            print(f"Failed to describe tables: {e}")
    
    def _cmd_connect(self, args):
        """Connect to a database"""
        if not args:
            print("Usage: \\c <database_name>")
            return
        
        db_name = args[0]
        try:
            print(f"Connecting to database '{db_name}'...")
            self._execute_query(f"USE {db_name}")
        except Exception as e:
            print(f"Connection failed: {e}")
    
    def _cmd_set_normal(self, args):
        """Set display mode to force normal (persistent)"""
        self.config.display_mode = 'force_normal'
        self.config.save_config()
        print("Display mode set to: force normal (persistent)")
        
        # Show last result if available
        if hasattr(self.config, 'last_result') and self.config.last_result:
            print("\nRe-displaying last result in normal format:")
            formatted_table = self.formatter.format_table(self.config.last_result)
            print(formatted_table)
    
    def _cmd_set_auto(self, args):
        """Set display mode to auto (persistent)"""
        self.config.display_mode = 'auto'
        self.config.save_config()
        print("Display mode set to: auto (persistent)")
        
        # Show last result if available
        if hasattr(self.config, 'last_result') and self.config.last_result:
            print("\nRe-displaying last result in auto format:")
            formatted_table = self.formatter.format_table(self.config.last_result)
            print(formatted_table)
    
    def _cmd_wide(self, args):
        """Show wide table in different formats"""
        if not hasattr(self.config, 'last_result') or not self.config.last_result:
            print("No query result available. Run a SELECT query first.")
            return
        
        if not args:
            print("Usage: \\wide [vertical]")
            return
        
        format_type = args[0].lower()
        columns = list(self.config.last_result[0].keys()) if self.config.last_result else []
        
        if format_type == 'vertical':
            result = self.formatter._format_wide_table(self.config.last_result, columns)
            print(result)
        else:
            print("Unknown format. Use 'vertical'")
    
    def _cmd_cols(self, args):
        """Show specific columns from last result"""
        if not hasattr(self.config, 'last_result') or not self.config.last_result:
            print("No query result available. Run a SELECT query first.")
            return
        
        if not args:
            # Show available columns
            columns = list(self.config.last_result[0].keys())
            print(f"Available columns ({len(columns)} total):")
            for i, col in enumerate(columns, 1):
                print(f"{i:3}. {col}")
            print("\nUsage: \\cols 1,3,5  or  \\cols name,age,email")
            return
        
        # Parse column specification
        col_spec = ' '.join(args)
        all_columns = list(self.config.last_result[0].keys())
        
        selected_columns = []
        for spec in col_spec.split(','):
            spec = spec.strip()
            if spec.isdigit():
                # Column number
                idx = int(spec) - 1
                if 0 <= idx < len(all_columns):
                    selected_columns.append(all_columns[idx])
            elif spec in all_columns:
                # Column name
                selected_columns.append(spec)
        
        if not selected_columns:
            print("No valid columns specified")
            return
        
        # Show selected columns
        formatted_table = self.formatter.format_table(self.config.last_result, selected_columns)
        print(formatted_table)
    
    def _cmd_show_csv_style(self, args):
        """Show table in CSV format for easy viewing"""
        if not hasattr(self.config, 'last_result') or not self.config.last_result:
            print("No query result available. Run a SELECT query first.")
            return
        
        print("CSV-style display:")
        
        columns = list(self.config.last_result[0].keys())
        
        # Header
        header = ",".join(columns)
        print(header)
        
        # Data rows
        for row in self.config.last_result:
            row_data = []
            for col in columns:
                value = self.formatter._format_value(row.get(col))
                # Escape commas and quotes for CSV
                if ',' in value or '"' in value:
                    value = f'"{value.replace('"', '""')}"'
                row_data.append(value)
            print(",".join(row_data))
    
    def _cmd_show_columns_only(self, args):
        from datetime import datetime, date, time
        """Show only the column names and types"""
        if not hasattr(self.config, 'last_result') or not self.config.last_result:
            print("No query result available. Run a SELECT query first.")
            return
        
        columns = list(self.config.last_result[0].keys())
        print(f"Column Names ({len(columns)} total):")
        
        for i, col in enumerate(columns, 1):
            # Try to determine type from first non-null value
            col_type = "UNKNOWN"
            for row in self.config.last_result:
                if row.get(col) is not None:
                    value = row.get(col)
                    if isinstance(value, int):
                        col_type = "INTEGER"
                    elif isinstance(value,date):
                        col_type = "DATE"
                    elif isinstance(value, time):
                        col_type = "TIME"
                    elif isinstance(value, datetime):
                        col_type = "TIMESTAMP"
                    elif isinstance(value, float):
                        col_type = "FLOAT"
                    elif isinstance(value, bool):
                        col_type = "BOOLEAN"
                    elif isinstance(value, str):
                        col_type = "TEXT"
                    break
            
            print(f"{i:3}. {col:30} ({col_type})")
    
    def _cmd_list_modules(self, args):
        """List loaded modules available for hot-reload"""
        print("Registered Modules:\n")
        
        if not MODULE_REGISTRY:
            print("No modules registered for hot-reload")
            return
        
        modules_info = []
        for name, module in MODULE_REGISTRY.items():
            try:
                module_file = getattr(module, '__file__', 'Unknown')
                module_time = datetime.fromtimestamp(os.path.getmtime(module_file)).strftime('%Y-%m-%d %H:%M:%S') if module_file != 'Unknown' else 'Unknown'
                
                modules_info.append({
                    'Module': name,
                    'File': os.path.basename(module_file) if module_file != 'Unknown' else 'Unknown',
                    'Modified': module_time
                })
            except Exception as e:
                modules_info.append({
                    'Module': name,
                    'File': 'Error',
                    'Modified': str(e)[:30]
                })
        
        formatted_table = self.formatter.format_table(modules_info, ['Module', 'File', 'Modified'])
        print(formatted_table)
    
    def _cmd_debug_mode(self, args):
        """Toggle debug mode"""
        # Toggle debug configuration
        debug_mode = not getattr(self.config, 'debug_mode', False)
        self.config.debug_mode = debug_mode
        
        if debug_mode:
            print("Debug mode enabled - detailed error reporting active")
            self.config.reload_on_error = True
        else:
            print("Debug mode disabled")
            self.config.reload_on_error = False
        
        self.config.save_config()
    
    def _cmd_toggle_timing(self, args):
        """Toggle query timing with visual feedback"""
        self.config.timing = not self.config.timing
        if self.config.timing:
            print("Query timing enabled")
        else:
            print("Query timing disabled")
        self.config.save_config()
    
    def _cmd_toggle_echo(self, args):
        """Toggle query echo"""
        self.config.echo_queries = not self.config.echo_queries
        if self.config.echo_queries:
            print("Query echo enabled")
        else:
            print("Query echo disabled")
        self.config.save_config()
    
    def _cmd_set_config(self, args):
        """Set configuration option with validation"""
        if len(args) < 2:
            print("Usage: \\set <option> <value>")
            print("Available options:")
            print("  table_format             ascii, markdown, csv")
            print("  max_column_width         Maximum width for table columns")
            print("  null_display             How to display NULL values")
            print("  auto_detect_wide         Enable/disable wide table detection (true/false)")
            print("  wide_table_threshold     Column count threshold for wide tables")
            print("  display_mode            'auto', 'force_normal', 'force_wide'")
            print("  auto_reload             Auto-reload modules on file changes")
            print("  reload_on_error         Auto-reload on execution errors")
            return
        
        option, value = args[0], args[1]
        
        try:
            if option == 'table_format' and value in ['ascii', 'markdown', 'csv']:
                self.config.table_format = value
                print(f"Table format set to: {value}")
            elif option == 'max_column_width':
                self.config.max_column_width = int(value)
                print(f"Max column width set to: {value}")
            elif option == 'null_display':
                self.config.null_display = value
                print(f"NULL display set to: '{value}'")
            elif option == 'auto_detect_wide':
                self.config.auto_detect_wide = value.lower() == 'true'
                status = "enabled" if self.config.auto_detect_wide else "disabled"
                print(f"Wide table auto-detection {status}")
            elif option == 'wide_table_threshold':
                self.config.wide_table_threshold = int(value)
                print(f"Wide table threshold set to {value} columns")
            elif option == 'display_mode' and value in ['auto', 'force_normal', 'force_wide']:
                self.config.display_mode = value
                print(f"Display mode set to: {value}")
            elif option == 'auto_reload':
                self.config.auto_reload = value.lower() == 'true'
                status = "enabled" if self.config.auto_reload else "disabled"
                print(f"Auto-reload {status}")
            elif option == 'reload_on_error':
                self.config.reload_on_error = value.lower() == 'true'
                status = "enabled" if self.config.reload_on_error else "disabled"
                print(f"Reload on error {status}")
            else:
                print(f"Unknown option: {option}")
                return
            
            self.config.save_config()
        
        except ValueError as e:
            print(f"Invalid value: {e}")
    
    def _cmd_show_config(self, args):
        """Show configuration"""
        if args:
            option = args[0]
            if hasattr(self.config, option):
                value = getattr(self.config, option)
                print(f"{option} = {value}")
            else:
                print(f"Unknown option: {option}")
        else:
            print("Current Configuration:")
            print(f"table_format = {self.config.table_format}")
            print(f"max_column_width = {self.config.max_column_width}")
            print(f"null_display = {self.config.null_display}")
            print(f"timing = {self.config.timing}")
            print(f"echo_queries = {self.config.echo_queries}")
            print(f"auto_detect_wide = {getattr(self.config, 'auto_detect_wide', False)}")
            print(f"wide_table_threshold = {getattr(self.config, 'wide_table_threshold', 10)}")
            print(f"display_mode = {getattr(self.config, 'display_mode', 'auto')}")
            print(f"auto_reload = {getattr(self.config, 'auto_reload', True)}")
            print(f"reload_on_error = {getattr(self.config, 'reload_on_error', False)}")
    
    def _cmd_show_history(self, args):
        """Show command history with enhanced display"""
        try:
            if PROMPT_TOOLKIT_AVAILABLE:
                # Get history from prompt_toolkit
                history_items = []
                for i, entry in enumerate(self.history.file_history.get_strings()):
                    history_items.append({'#': i+1, 'Command': entry[:80] + ('...' if len(entry) > 80 else '')})
                
                if history_items:
                    print("Command History:\n")
                    # Show last 20 items
                    recent_items = history_items[-20:] if len(history_items) > 20 else history_items
                    formatted_table = self.formatter.format_table(recent_items, ['#', 'Command'])
                    print(formatted_table)
                    
                    if len(history_items) > 20:
                        print(f"\nShowing last 20 of {len(history_items)} commands")
                else:
                    print("No command history available")
            else:
                # Fallback to readline history
                try:
                    import readline
                    print("Command History:\n")
                    for i in range(max(0, readline.get_current_history_length() - 20), readline.get_current_history_length()):
                        line = readline.get_history_item(i + 1)
                        if line:
                            print(f"{i+1:4}: {line}")
                except:
                    print("History not available")
        except Exception as e:
            print(f"Could not retrieve history: {e}")
    
    def _cmd_clear(self, args):
        """Clear the screen"""
        os.system('cls' if os.name == 'nt' else 'clear')
        # Reprint a minimal banner
        
    
    def _cmd_export(self, args):
        """Export last query result with enhanced options"""
        if len(args) < 2:
            print("Usage: \\export <format> <filename>")
            print("Supported formats: csv, json, sql, xlsx")
            return
        
        if not self.config.last_result:
            print("No query result to export. Run a SELECT query first.")
            return
        
        format_type = args[0].lower()
        filename = args[1]
        
        try:
            print(f"Exporting {len(self.config.last_result)} rows to {filename}...")
            
            if format_type == 'csv':
                self._export_csv(filename)
            elif format_type == 'json':
                self._export_json(filename)
            elif format_type == 'sql':
                self._export_sql(filename)
            elif format_type == 'xlsx':
                self._export_xlsx(filename)
            else:
                print(f"Unsupported format: {format_type}")
                print("Supported formats: csv, json, sql, xlsx")
                return
            
            row_count = len(self.config.last_result)
            file_size = os.path.getsize(filename) if os.path.exists(filename) else 0
            size_str = f"{file_size:,} bytes" if file_size < 1024 else f"{file_size//1024:,} KB"
            
            print(f"Exported {row_count:,} rows to {filename} ({format_type.upper()}, {size_str})")
            
        except Exception as e:
            print(f"Export failed: {str(e)}")
    
    def _export_csv(self, filename):
        """Export to CSV format"""
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            if not self.config.last_result:
                return
            
            columns = list(self.config.last_result[0].keys())
            writer = csv.writer(csvfile)
            
            writer.writerow(columns)
            
            for row in self.config.last_result:
                row_data = [row.get(col, '') for col in columns]
                writer.writerow(row_data)
    
    def _export_json(self, filename):
        """Export to JSON format"""
        if not filename.endswith('.json'):
            filename += '.json'
        
        with open(filename, 'w', encoding='utf-8') as jsonfile:
            json.dump({
                'exported_at': datetime.now().isoformat(),
                'row_count': len(self.config.last_result),
                'data': self.config.last_result
            }, jsonfile, indent=2, default=str)
    
    def _export_sql(self, filename):
        """Export as SQL INSERT statements"""
        if not filename.endswith('.sql'):
            filename += '.sql'
        
        if not self.config.last_result:
            return
        
        # Ask for table name with a default
        try:
            if PROMPT_TOOLKIT_AVAILABLE:
                table_name = prompt("Enter table name for INSERT statements (default: exported_data): ").strip()
            else:
                table_name = input("Enter table name for INSERT statements (default: exported_data): ").strip()
        except (KeyboardInterrupt, EOFError):
            table_name = ""
        
        if not table_name:
            table_name = "exported_data"
        
        with open(filename, 'w', encoding='utf-8') as sqlfile:
            columns = list(self.config.last_result[0].keys())
            
            # Write header
            sqlfile.write(f"-- MyShell SQL Export\n")
            sqlfile.write(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            sqlfile.write(f"-- Rows: {len(self.config.last_result):,}\n")
            sqlfile.write(f"-- Table: {table_name}\n\n")
            
            # Write INSERT statements
            for i, row in enumerate(self.config.last_result):
                values = []
                for col in columns:
                    value = row.get(col)
                    if value is None:
                        values.append('NULL')
                    elif isinstance(value, str):
                        escaped_value = value.replace("'", "''")
                        values.append(f"'{escaped_value}'")
                    elif isinstance(value, bool):
                        values.append('TRUE' if value else 'FALSE')
                    else:
                        values.append(str(value))
                
                columns_str = ', '.join(columns)
                values_str = ', '.join(values)
                sqlfile.write(f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});\n")
                
                # Add progress comments for large exports
                if i > 0 and i % 1000 == 0:
                    sqlfile.write(f"\n-- Progress: {i:,} rows inserted\n")
    
    def _export_xlsx(self, filename):
        """Export to Excel format (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("Excel export requires openpyxl: pip install openpyxl")
            return
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyShell Export"
        
        if not self.config.last_result:
            return
        
        columns = list(self.config.last_result[0].keys())
        
        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, row in enumerate(self.config.last_result, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def _cmd_import(self, args):
        """Import SQL file with progress tracking"""
        if not args:
            print("Usage: \\import <filename>")
            return
        
        filename = args[0]
        
        if not os.path.exists(filename):
            print(f"File not found: {filename}")
            return
        
        try:
            print(f"Importing SQL from {filename}...")
            
            with open(filename, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Split into statements
            statements = [s.strip() for s in content.split(';') if s.strip()]
            
            if not statements:
                print("No SQL statements found in file")
                return
            
            print(f"Found {len(statements)} statements to execute...")
            
            # Execute statements with progress
            successful = 0
            failed = 0
            
            for i, stmt in enumerate(statements, 1):
                try:
                    if len(stmt) > 50:
                        preview = stmt[:50] + "..."
                    else:
                        preview = stmt
                    
                    print(f"[{i:3}/{len(statements)}] {preview}")
                    self._execute_query(stmt + ';')
                    successful += 1
                    
                except Exception as e:
                    print(f"Statement {i} failed: {e}")
                    failed += 1
                    
                    # Ask user if they want to continue
                    if failed >= 3:
                        try:
                            if PROMPT_TOOLKIT_AVAILABLE:
                                continue_import = prompt("Multiple failures detected. Continue? (y/n): ").lower()
                            else:
                                continue_import = input("Multiple failures detected. Continue? (y/n): ").lower()
                            
                            if continue_import not in ['y', 'yes']:
                                break
                        except (KeyboardInterrupt, EOFError):
                            break
            
            # Summary
            total = successful + failed
            if successful > 0:
                print(f"Import completed: {successful}/{total} statements successful")
            if failed > 0:
                print(f"{failed} statements failed")
                
        except Exception as e:
            print(f"Import failed: {e}")
    
    def _cmd_version(self, args):
        """Show comprehensive version information"""
        python_version = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        
        version_info = f"""

MySQL Shell CLI
  Version: 2.1.0
  
Runtime Environment
  Python: {python_version}
  Platform: {sys.platform}
  
Features
  Hot-reload: Enabled
  Enhanced UI: {'Available' if PROMPT_TOOLKIT_AVAILABLE else 'Missing (pip install prompt-toolkit)'}
  Export formats: CSV, JSON, SQL, XLSX
  Multi-query support: Enabled
  Comment support: Enabled (-- syntax)
  
Modules Status"""

        # Check module status
        module_status = []
        for name in ['database_manager', 'engine', 'executor', 'datatypes']:
            if name in MODULE_REGISTRY:
                module_status.append(f"  {name}: Loaded")
            else:
                module_status.append(f"  {name}: Missing")
        
        version_info += '\n' + '\n'.join(module_status)
        
        # Add session info
        uptime = datetime.now() - self.start_time
        version_info += f"""

Session Info
  Uptime: {uptime}
  Queries executed: {self.query_count:,}
  Display mode: {getattr(self.config, 'display_mode', 'auto')}
"""
        
        print(version_info)
    
    def _cmd_status(self, args):
        """Show comprehensive session and connection status"""
        uptime = datetime.now() - self.start_time
        
        # Get database info
        try:
            current_db = getattr(db_manager, 'active_db_name', None)
            if current_db:
                db_name = os.path.basename(current_db)
                if db_name.endswith('.su'):
                    db_name = db_name[:-3]
                db_status = db_name
                
                # Get table count
                try:
                    table_count = len(db_manager.active_db)
                    db_info = f" ({table_count} tables)"
                except:
                    db_info = ""
            else:
                db_status = "Not connected"
                db_info = ""
        except:
            db_status = "Database manager unavailable"
            db_info = ""
        
        # Memory usage (if available)
        try:
            import psutil
            process = psutil.Process()
            memory_mb = process.memory_info().rss / 1024 / 1024
            memory_info = f"{memory_mb:.1f} MB"
        except:
            memory_info = "N/A"
        
        status_info = f"""

Database Connection
  Current database: {db_status}{db_info}
  Available databases: {len(getattr(db_manager, 'databases', []))}

Session Statistics
  Queries executed: {self.query_count:,}
  Session uptime: {uptime}
  Memory usage: {memory_info}

Configuration
  Query timing: {'ON' if self.config.timing else 'OFF'}
  Query echo: {'ON' if self.config.echo_queries else 'OFF'}
  Table format: {self.config.table_format}
  Display mode: {getattr(self.config, 'display_mode', 'auto')}
  Auto-reload: {'ON' if getattr(self.config, 'auto_reload', True) else 'OFF'}
  Wide table detection: {'ON' if getattr(self.config, 'auto_detect_wide', False) else 'OFF'}
  
Hot-Reload Status
  Registered modules: {len(MODULE_REGISTRY)}
  Last result rows: {len(self.config.last_result) if self.config.last_result else 0:,}
"""
        print(status_info)


def main():
    """Enhanced main entry point with better argument handling"""
    os.system('cls' if os.name == 'nt' else 'clear')
    
    parser = argparse.ArgumentParser(
        
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  myshell                          # Start interactive shell
  myshell --execute "SELECT * FROM users;"  # Execute query and exit
  myshell --config /path/config.json       # Use custom config file
  
Hot-reload commands:
  \\modules                        # List loaded modules
  
Multi-query support:
  INSERT INTO users VALUES (1, 'John'); INSERT INTO users VALUES (2, 'Jane');
  Use Ctrl+Enter for new lines within queries
        """
    )
    
    parser.add_argument('--config', 
                       help='Path to config file')
    parser.add_argument('--execute', '-e', 
                       help='Execute command and exit')
    parser.add_argument('--version', action='version', version='MyShell SQL CLI 2.1.0')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug mode with detailed error reporting')
    parser.add_argument('--import-file', 
                       help='Import SQL file on startup')
    parser.add_argument('--export-config', 
                       help='Export current configuration to file')
    
    args = parser.parse_args()
    
    # Handle early exit operations
    if args.export_config:
        try:
            shell = EnhancedSQLShell()
            shell.config.save_config()
            print(f"Configuration exported to {shell.config.config_file}")
            return
        except Exception as e:
            print(f"Failed to export config: {e}")
            sys.exit(1)
    
    try:
        shell = EnhancedSQLShell()
        
        # Apply command line options
        if args.debug:
            shell.config.debug_mode = True
            shell.config.reload_on_error = True
            print("Debug mode enabled")
        
        if args.config:
            # Load custom config file
            try:
                shell.config.config_file = Path(args.config)
                shell.config.load_config()
                print(f"Loaded config from {args.config}")
            except Exception as e:
                print(f"Failed to load config: {e}")
        
        # Handle one-time operations
        if args.execute:
            shell._execute_query(args.execute)
            return
        
        if args.import_file:
            shell._cmd_import([args.import_file])
            if not sys.stdin.isatty():  # If running in script mode
                return
        
        # Start interactive shell
        shell.run()
        
    except KeyboardInterrupt:
        print("\nInterrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"Fatal error: {e}")
        if args.debug:
            print(f"{traceback.format_exc()}")
        sys.exit(1)


if __name__ == '__main__':
    main()