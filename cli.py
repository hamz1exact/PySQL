#!/usr/bin/env python3
"""
MySQL Shell - A professional command-line interface for your SQL engine
Enhanced version with hot-reloading and better terminal UI
"""

import sys
import os
import time
import json
import csv
import importlib
import atexit
import re
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import argparse

# Enhanced terminal UI imports
try:
    from prompt_toolkit import prompt
    from prompt_toolkit.completion import WordCompleter
    from prompt_toolkit.history import FileHistory
    from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
    from prompt_toolkit.shortcuts import print_formatted_text
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.styles import Style
    PROMPT_TOOLKIT_AVAILABLE = True
except ImportError:
    PROMPT_TOOLKIT_AVAILABLE = False
    print("Warning: prompt_toolkit not available. Install with: pip install prompt-toolkit")

# Module registry for hot-reloading
MODULE_REGISTRY = {}

def register_module(name, module):
    """Register a module for hot-reloading"""
    MODULE_REGISTRY[name] = module

def reload_modules():
    """Hot-reload all registered modules"""
    reloaded = []
    failed = []
    
    for name, module in MODULE_REGISTRY.items():
        try:
            importlib.reload(module)
            reloaded.append(name)
        except Exception as e:
            failed.append((name, str(e)))
    
    return reloaded, failed

# Try to import your existing modules
try:
    import database_manager
    import engine
    import executor
    import datatypes
    
    # Register modules for hot-reloading
    register_module('database_manager', database_manager)
    register_module('engine', engine)
    register_module('executor', executor)
    register_module('datatypes', datatypes)
    
    from engine import db_manager, Lexer, Parser
    from executor import execute
    
except ImportError as e:
    print(f"Warning: Could not import SQL engine modules: {e}")
    print("Make sure your engine.py, executor.py, and database_manager.py are in the Python path")
    
    # Create mock objects for testing
    class MockDBManager:
        def __init__(self):
            self.active_db = {}
            self.active_db_name = None
            self.databases = []
        
        def save_database_file(self):
            pass
    
    class MockLexer:
        def __init__(self, query):
            self.tokens = [("SELECT", "SELECT")]
    
    class MockParser:
        def __init__(self, tokens):
            self.tokens = tokens
        
        def parse_select_statement(self):
            return {"type": "SELECT", "mock": True}
    
    def mock_execute(ast, database):
        return [{"message": "Mock execution - engine modules not available"}]
    
    db_manager = MockDBManager()
    Lexer = MockLexer
    Parser = MockParser
    execute = mock_execute


class Colors:
    """Enhanced ANSI color codes for terminal output"""
    RESET = '\033[0m'
    BOLD = '\033[1m'
    DIM = '\033[2m'
    ITALIC = '\033[3m'
    UNDERLINE = '\033[4m'
    
    # Foreground colors
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    
    # Bright colors
    BRIGHT_BLACK = '\033[90m'
    BRIGHT_RED = '\033[91m'
    BRIGHT_GREEN = '\033[92m'
    BRIGHT_YELLOW = '\033[93m'
    BRIGHT_BLUE = '\033[94m'
    BRIGHT_MAGENTA = '\033[95m'
    BRIGHT_CYAN = '\033[96m'
    BRIGHT_WHITE = '\033[97m'


class Config:
    """Enhanced configuration with hot-reload support"""
    def __init__(self):
        self.show_query_time = True
        self.show_row_count = True
        self.max_column_width = 50
        self.null_display = 'NULL'
        self.prompt_format = '{db}> '
        self.pager = 'less'
        self.history_size = 1000
        self.auto_commit = True
        self.echo_queries = False
        self.colorize_output = True
        self.table_format = 'ascii'  # ascii, markdown, csv
        self.timing = True
        self.last_result = None
        self.last_columns = None
        self.auto_reload = True  # Enable auto-reload on file changes
        self.reload_on_error = False  # Reload modules on execution error
        
        # Load config from file if exists
        self.config_file = Path.home() / '.myshell_config.json'
        self.load_config()
    
    def load_config(self):
        """Load configuration from file"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config_data = json.load(f)
                    for key, value in config_data.items():
                        if hasattr(self, key):
                            setattr(self, key, value)
        except Exception as e:
            print(f"Warning: Could not load config file: {e}")
    
    def save_config(self):
        """Save configuration to file"""
        try:
            config_data = {
                'show_query_time': self.show_query_time,
                'show_row_count': self.show_row_count,
                'max_column_width': self.max_column_width,
                'null_display': self.null_display,
                'prompt_format': self.prompt_format,
                'pager': self.pager,
                'history_size': self.history_size,
                'auto_commit': self.auto_commit,
                'echo_queries': self.echo_queries,
                'colorize_output': self.colorize_output,
                'table_format': self.table_format,
                'timing': self.timing,
                'auto_reload': self.auto_reload,
                'reload_on_error': self.reload_on_error
            }
            with open(self.config_file, 'w') as f:
                json.dump(config_data, f, indent=2)
        except Exception as e:
            print(f"Warning: Could not save config file: {e}")


class EnhancedHistoryManager:
    """Enhanced history management with prompt_toolkit support"""
    def __init__(self, config: Config):
        self.config = config
        self.history_file = Path.home() / '.myshell_history'
        
        if PROMPT_TOOLKIT_AVAILABLE:
            self.file_history = FileHistory(str(self.history_file))
        else:
            self.setup_readline_history()
    
    def setup_readline_history(self):
        """Fallback to readline history"""
        try:
            import readline
            readline.set_history_length(self.config.history_size)
            
            if self.history_file.exists():
                readline.read_history_file(str(self.history_file))
            
            atexit.register(self.save_readline_history)
        except ImportError:
            print("Warning: readline not available, history disabled")
    
    def save_readline_history(self):
        """Save readline history"""
        try:
            import readline
            readline.write_history_file(str(self.history_file))
        except:
            pass


class TableFormatter:
    """Enhanced table formatting with better visual appeal"""
    
    def __init__(self, config: Config):
        self.config = config
    
    def format_table(self, rows: List[Dict], columns: List[str] = None) -> str:
        """Format table data for display"""
        if not rows:
            return self._colorize("Empty result set", Colors.BRIGHT_BLACK)
        
        if columns is None:
            columns = list(rows[0].keys()) if rows else []
        
        if self.config.table_format == 'csv':
            return self._format_csv(rows, columns)
        elif self.config.table_format == 'markdown':
            return self._format_markdown(rows, columns)
        else:
            return self._format_ascii_enhanced(rows, columns)
    
    def _format_ascii_enhanced(self, rows: List[Dict], columns: List[str]) -> str:
        """Enhanced ASCII table with better borders and colors"""
        if not rows:
            return self._colorize("Empty result set", Colors.BRIGHT_BLACK)
        
        # Calculate column widths
        col_widths = {}
        for col in columns:
            col_widths[col] = min(len(col), self.config.max_column_width)
            
        for row in rows:
            for col in columns:
                value_str = self._format_value(row.get(col))
                col_widths[col] = max(col_widths[col], min(len(value_str), self.config.max_column_width))
        
        lines = []
        
        # Enhanced borders with double lines for header
        def make_border(char='─'):
            parts = []
            for col in columns:
                parts.append(char * (col_widths[col] + 2))
            return '┌' + '┬'.join(parts) + '┐' if char == '─' else '├' + '┼'.join(parts) + '┤'
        
        # Top border
        lines.append(make_border())
        
        # Header with enhanced styling
        header_parts = []
        for col in columns:
            header_text = col[:col_widths[col]].ljust(col_widths[col])
            if self.config.colorize_output:
                header_text = self._colorize(header_text, Colors.BOLD + Colors.BRIGHT_CYAN)
            header_parts.append(f' {header_text} ')
        lines.append('│' + '│'.join(header_parts) + '│')
        
        # Header separator
        lines.append(make_border('─'))
        
        # Data rows with alternating colors
        for i, row in enumerate(rows):
            row_parts = []
            for col in columns:
                value = row.get(col)
                value_str = self._format_value(value)
                
                if len(value_str) > col_widths[col]:
                    value_str = value_str[:col_widths[col]-3] + '...'
                
                value_str = value_str.ljust(col_widths[col])
                
                if self.config.colorize_output:
                    # Alternating row colors for better readability
                    if i % 2 == 0:
                        value_str = self._colorize_value(value_str, value)
                    else:
                        value_str = self._colorize_value(value_str, value, dim=True)
                
                row_parts.append(f' {value_str} ')
            lines.append('│' + '│'.join(row_parts) + '│')
        
        # Bottom border
        bottom_parts = []
        for col in columns:
            bottom_parts.append('─' * (col_widths[col] + 2))
        lines.append('└' + '┴'.join(bottom_parts) + '┘')
        
        return '\n'.join(lines)
    
    def _format_markdown(self, rows: List[Dict], columns: List[str]) -> str:
        """Enhanced Markdown table format"""
        if not rows:
            return "| Empty result set |"
        
        lines = []
        
        # Header with better alignment
        header = '| ' + ' | '.join(f"**{col}**" for col in columns) + ' |'
        lines.append(header)
        
        # Enhanced separator
        separator = '| ' + ' | '.join([':---:'] * len(columns)) + ' |'
        lines.append(separator)
        
        # Data rows
        for row in rows:
            row_data = []
            for col in columns:
                value = self._format_value(row.get(col))
                # Escape pipe characters in markdown
                value = value.replace('|', '\\|')
                row_data.append(value)
            lines.append('| ' + ' | '.join(row_data) + ' |')
        
        return '\n'.join(lines)
    
    def _format_csv(self, rows: List[Dict], columns: List[str]) -> str:
        """CSV format implementation"""
        lines = []
        lines.append(','.join(columns))
        
        for row in rows:
            row_data = []
            for col in columns:
                value = self._format_value(row.get(col))
                if ',' in value or '"' in value or '\n' in value:
                    value = '"' + value.replace('"', '""') + '"'
                row_data.append(value)
            lines.append(','.join(row_data))
        
        return '\n'.join(lines)
    
    def _format_value(self, value) -> str:
        """Format a single value for display"""
        if value is None:
            return self.config.null_display
        elif isinstance(value, bool):
            return 'TRUE' if value else 'FALSE'
        elif isinstance(value, (int, float)):
            return str(value)
        else:
            return str(value)
    
    def _colorize_value(self, text: str, value, dim=False) -> str:
        """Apply color based on value type with optional dimming"""
        base_color = ""
        if value is None:
            base_color = Colors.BRIGHT_BLACK
        elif isinstance(value, bool):
            base_color = Colors.BRIGHT_MAGENTA
        elif isinstance(value, (int, float)):
            base_color = Colors.BRIGHT_GREEN
        elif isinstance(value, str):
            base_color = Colors.WHITE
        
        if dim and base_color:
            base_color = Colors.DIM + base_color
        
        return self._colorize(text, base_color) if base_color else text
    
    def _colorize(self, text: str, color: str) -> str:
        """Apply color to text"""
        if self.config.colorize_output:
            return f"{color}{text}{Colors.RESET}"
        return text


class EnhancedSQLShell:
    """Enhanced SQL shell with hot-reloading and better terminal UI"""
    
    def __init__(self):
        self.config = Config()
        self.history = EnhancedHistoryManager(self.config)
        self.formatter = TableFormatter(self.config)
        self.query_count = 0
        self.start_time = datetime.now()
        
        # SQL keywords for completion
        self.sql_keywords = [
            'SELECT', 'FROM', 'WHERE', 'INSERT', 'UPDATE', 'DELETE',
            'CREATE', 'DROP', 'ALTER', 'TABLE', 'DATABASE', 'INDEX',
            'AND', 'OR', 'NOT', 'NULL', 'IS', 'IN', 'LIKE', 'BETWEEN',
            'ORDER', 'BY', 'GROUP', 'HAVING', 'DISTINCT', 'COUNT',
            'SUM', 'AVG', 'MAX', 'MIN', 'JOIN', 'LEFT', 'RIGHT',
            'INNER', 'OUTER', 'ON', 'AS', 'USE'
        ]
        
        # Setup completions
        if PROMPT_TOOLKIT_AVAILABLE:
            self.completer = WordCompleter(self.sql_keywords, ignore_case=True)
            self.style = Style.from_dict({
                'prompt': '#ansibrightblue bold',
                'continuation': '#ansibrightblack',
                'error': '#ansibrightred',
                'success': '#ansibrightgreen',
                'warning': '#ansiyellow',
                'info': '#ansibrightcyan',
            })
        
        # Enhanced command mappings
        self.meta_commands = {
            '\\q': self._cmd_quit,
            '\\quit': self._cmd_quit,
            '\\exit': self._cmd_quit,
            '\\h': self._cmd_help,
            '\\help': self._cmd_help,
            '\\?': self._cmd_help,
            '\\l': self._cmd_list_databases,
            '\\list': self._cmd_list_databases,
            '\\d': self._cmd_describe_tables,
            '\\dt': self._cmd_describe_tables,
            '\\c': self._cmd_connect,
            '\\connect': self._cmd_connect,
            '\\use': self._cmd_connect,
            '\\timing': self._cmd_toggle_timing,
            '\\echo': self._cmd_toggle_echo,
            '\\set': self._cmd_set_config,
            '\\show': self._cmd_show_config,
            '\\history': self._cmd_show_history,
            '\\clear': self._cmd_clear,
            '\\cls': self._cmd_clear,
            '\\export': self._cmd_export,
            '\\import': self._cmd_import,
            '\\version': self._cmd_version,
            '\\status': self._cmd_status,
            '\\r': self._cmd_reload,
            '\\reload': self._cmd_reload,
            '\\refresh': self._cmd_reload,
            '\\modules': self._cmd_list_modules,
            '\\debug': self._cmd_debug_mode,
        }
    

        

    
    def _get_connection_info(self) -> str:
        """Get current connection information"""
        try:
            current_db = getattr(db_manager, 'active_db_name', None)
            if current_db:
                filename = os.path.basename(current_db)
                if filename.endswith('.su'):
                    filename = filename[:-3]
                return f"{Colors.GREEN}Connected to: {Colors.BOLD}{filename}{Colors.RESET}"
            else:
                return f"{Colors.YELLOW}No database selected. Use {Colors.BOLD}USE database_name{Colors.RESET}{Colors.YELLOW} to connect.{Colors.RESET}"
        except:
            return f"{Colors.RED}Database manager not available{Colors.RESET}"
    
    def get_prompt_text(self) -> str:
        """Generate the command prompt text"""
        try:
            db_name = getattr(db_manager, 'active_db_name', None) or 'myshell'
            
            if db_name and db_name != 'myshell':
                filename = os.path.basename(db_name)
                if filename.endswith('.su'):
                    filename = filename[:-3]
                db_name = filename
        except:
            db_name = 'myshell'
        
        return f"{db_name}> "
    
    def get_formatted_prompt(self):
        """Get formatted prompt for prompt_toolkit"""
        if PROMPT_TOOLKIT_AVAILABLE:
            db_name = 'myshell'
            try:
                current_db = getattr(db_manager, 'active_db_name', None)
                if current_db:
                    filename = os.path.basename(current_db)
                    if filename.endswith('.su'):
                        filename = filename[:-3]
                    db_name = filename
            except:
                pass
            
            return FormattedText([
                ('class:prompt', f"{db_name}> "),
            ])
        else:
            return self.get_prompt_text()
    
    def run(self):
        """Enhanced main shell loop with better error handling"""
        
        query_buffer = []
        
        while True:
            try:
                # Get input with enhanced prompting
                if PROMPT_TOOLKIT_AVAILABLE:
                    if query_buffer:
                        continuation_prompt = FormattedText([('class:continuation', '    -> ')])
                        line = prompt(continuation_prompt, 
                                    style=self.style,
                                    history=self.history.file_history,
                                    auto_suggest=AutoSuggestFromHistory(),
                                    completer=self.completer).strip()
                    else:
                        line = prompt(self.get_formatted_prompt(),
                                    style=self.style,
                                    history=self.history.file_history,
                                    auto_suggest=AutoSuggestFromHistory(),
                                    completer=self.completer).strip()
                else:
                    # Fallback to basic input
                    if query_buffer:
                        prompt_text = f"{Colors.DIM}    -> {Colors.RESET}" if self.config.colorize_output else "    -> "
                    else:
                        prompt_text = self.get_prompt_text()
                    line = input(prompt_text).strip()
                
                if not line:
                    continue
                
                # Handle meta-commands
                if line.startswith('\\'):
                    if query_buffer:
                        self._print_warning("Discarding incomplete query")
                        query_buffer = []
                    
                    self._handle_meta_command(line)
                    continue
                
                # Build query buffer
                query_buffer.append(line)
                
                # Execute when query is complete
                full_query = ' '.join(query_buffer)
                if full_query.rstrip().endswith(';'):
                    self._execute_query(full_query)
                    query_buffer = []
                
            except KeyboardInterrupt:
                self._print_warning("^C - Query cancelled")
                query_buffer = []
                continue
            except EOFError:
                print(f"\n{Colors.BRIGHT_BLACK}Goodbye!{Colors.RESET}")
                break
            except Exception as e:
                self._print_error(f"Unexpected error: {e}")
                if self.config.reload_on_error:
                    self._print_info("Auto-reloading modules due to error...")
                    self._cmd_reload([])
    
    def _print_success(self, message):
        """Print success message"""
        if PROMPT_TOOLKIT_AVAILABLE:
            print_formatted_text(FormattedText([('class:success', f"{message}")]))
        else:
            print(f"{Colors.BRIGHT_GREEN}{message}{Colors.RESET}")
    
    def _print_error(self, message):
        """Print error message"""
        if PROMPT_TOOLKIT_AVAILABLE:
            print_formatted_text(FormattedText([('class:error', f"{message}")]))
        else:
            print(f"{Colors.BRIGHT_RED}{message}{Colors.RESET}")
    
    def _print_warning(self, message):
        """Print warning message"""
        if PROMPT_TOOLKIT_AVAILABLE:
            print_formatted_text(FormattedText([('class:warning', f"{message}")]))
        else:
            print(f"{Colors.YELLOW}{message}{Colors.RESET}")
    
    def _print_info(self, message):
        """Print info message"""
        if PROMPT_TOOLKIT_AVAILABLE:
            print_formatted_text(FormattedText([('class:info', f"{message}")]))
        else:
            print(f"{Colors.BRIGHT_CYAN}{message}{Colors.RESET}")
    
    def _handle_meta_command(self, command: str):
        """Handle meta-commands with enhanced error reporting"""
        parts = command.split()
        cmd = parts[0].lower()
        args = parts[1:] if len(parts) > 1 else []
        
        if cmd in self.meta_commands:
            try:
                self.meta_commands[cmd](args)
            except Exception as e:
                self._print_error(f"Command failed: {e}")
                print(f"{Colors.BRIGHT_BLACK}Details: {traceback.format_exc()}{Colors.RESET}")
        else:
            self._print_error(f"Unknown command: {cmd}")
            print(f"Type {Colors.BOLD}\\h{Colors.RESET} for help")
    
    def _execute_query(self, query: str):
        """Execute SQL query with enhanced error handling and hot-reload support"""
        query = query.strip()
        if not query:
            return
        
        if self.config.echo_queries:
            print(f"{Colors.DIM}{query}{Colors.RESET}")
        
        start_time = time.time()
        
        try:
            # Get current database
            database = db_manager.active_db
            
            # Use your existing Lexer and Parser
            lexer = Lexer(query)
            parser = Parser(lexer.tokens)
            
            if not lexer.tokens:
                return
            
            token_type = lexer.tokens[0][0]
            next_token_type = lexer.tokens[1][0] if len(lexer.tokens) > 1 else None
            
            # Parse and execute based on query type
            if token_type == "SELECT":
                ast = parser.parse_select_statement()
                result = execute(ast, database)
                self._handle_select_result(result, start_time)
                
            elif token_type == "INSERT":
                ast = parser.parse_insert_statement()
                result = execute(ast, database)
                db_manager.save_database_file()
                self._handle_modify_result("INSERT", start_time)
                
            elif token_type == "UPDATE":
                ast = parser.parse_update_statement()
                result = execute(ast, database)
                db_manager.save_database_file()
                self._handle_modify_result("UPDATE", start_time)
                
            elif token_type == "DELETE":
                ast = parser.parse_delete_statement()
                result = execute(ast, database)
                db_manager.save_database_file()
                self._handle_modify_result("DELETE", start_time)
                
            elif token_type == "CREATE":
                if next_token_type == "DATABASE":
                    ast = parser.parse_create_database()
                    result = execute(ast, db_manager)
                    self._handle_ddl_result("CREATE DATABASE", start_time)
                elif next_token_type == "TABLE":
                    ast = parser.parse_create_table()
                    execute(ast, db_manager)
                    db_manager.save_database_file()
                    self._handle_ddl_result("CREATE TABLE", start_time)
                else:
                    self._print_error("Unsupported CREATE statement")
                    
            elif token_type == "DROP":
                if next_token_type == "DATABASE":
                    ast = parser.parse_drop_database()
                    result = execute(ast, database)
                    self._handle_ddl_result("DROP DATABASE", start_time)
                elif next_token_type == "TABLE":
                    ast = parser.parse_drop_table()
                    result = execute(ast, database)
                    db_manager.save_database_file()
                    self._handle_ddl_result("DROP TABLE", start_time)
                else:
                    self._print_error("Unsupported DROP statement")
                    
            elif token_type == "ALTER":
                ast = parser.parse_alter_statement()
                result = execute(ast, database)
                db_manager.save_database_file()
                self._handle_ddl_result("ALTER TABLE", start_time)
                
            elif token_type == "USE":
                ast = parser.parse_use_statement()
                result = execute(ast, db_manager)
                self._handle_use_result("USE", start_time)
                
            else:
                self._print_error(f"Unsupported statement type '{token_type}'")
            
            self.query_count += 1
            
        except Exception as e:
            execution_time = time.time() - start_time
            self._print_error(f"Query execution failed: {str(e)}")
            
            if self.config.timing:
                time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
                print(f"{Colors.BRIGHT_BLACK}Time: {time_str}{Colors.RESET}")
            
            # Auto-reload on error if enabled
            if self.config.reload_on_error:
                self._print_info("Auto-reloading modules...")
                self._cmd_reload([])
        
        print()  # Empty line for readability
    
    def _handle_select_result(self, result, start_time):
        """Handle SELECT query results with enhanced display"""
        execution_time = time.time() - start_time
        
        # Store result for export functionality
        self.config.last_result = result
        
        if result:
            # Format and display table
            formatted_table = self.formatter.format_table(result)
            print(formatted_table)
            
            # Enhanced summary
            if self.config.show_row_count:
                row_count = len(result)
                plural = 's' if row_count != 1 else ''
                summary = f"({row_count} row{plural} returned)"
                if self.config.colorize_output:
                    print(f"\n{Colors.BRIGHT_BLACK}{summary}{Colors.RESET}")
                else:
                    print(f"\n{summary}")
        else:
            self._print_info("Query returned no results")
            self.config.last_result = None
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            timing_msg = f"Execution time: {time_str}"
            if self.config.colorize_output:
                print(f"{Colors.BRIGHT_BLACK}{timing_msg}{Colors.RESET}")
            else:
                print(timing_msg)
    
    def _handle_modify_result(self, operation, start_time):
        """Handle INSERT/UPDATE/DELETE results"""
        execution_time = time.time() - start_time
        self._print_success(f"{operation} operation completed")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"{Colors.BRIGHT_BLACK}Time: {time_str}{Colors.RESET}")
    
    def _handle_ddl_result(self, operation, start_time):
        """Handle CREATE/DROP/ALTER results"""
        execution_time = time.time() - start_time
        self._print_success(f"{operation} completed successfully")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"{Colors.BRIGHT_BLACK}Time: {time_str}{Colors.RESET}")
    
    def _handle_use_result(self, operation, start_time):
        """Handle USE database results"""
        execution_time = time.time() - start_time
        
        try:
            current_db_name = getattr(db_manager, 'active_db_name', 'unknown')
            if current_db_name and current_db_name != 'unknown':
                filename = os.path.basename(current_db_name)
                if filename.endswith('.su'):
                    filename = filename[:-3]
                current_db_name = filename
            
            self._print_success(f"Switched to database: {current_db_name}")
        except:
            self._print_success("Database switch completed")
        
        if self.config.timing:
            time_str = f"{execution_time:.3f}s" if execution_time < 1 else f"{execution_time:.2f}s"
            print(f"{Colors.BRIGHT_BLACK}Time: {time_str}{Colors.RESET}")
    
    # Enhanced Meta-command implementations
    def _cmd_quit(self, args):
        """Quit the shell with cleanup"""
        self.config.save_config()
        print(f"{Colors.BRIGHT_GREEN}Session lasted {datetime.now() - self.start_time}{Colors.RESET}")
        sys.exit(0)
    
    def _cmd_help(self, args):
        """Show comprehensive help information"""
        help_text = f"""


{Colors.BOLD}Meta Commands:{Colors.RESET}
  {Colors.BRIGHT_BLUE}\\h, \\help{Colors.RESET}              Show this help
  {Colors.BRIGHT_BLUE}\\q, \\quit, \\exit{Colors.RESET}      Quit the shell
  {Colors.BRIGHT_BLUE}\\l, \\list{Colors.RESET}              List all databases
  {Colors.BRIGHT_BLUE}\\d, \\dt{Colors.RESET}                List tables in current database
  {Colors.BRIGHT_BLUE}\\c <db>, \\connect <db>{Colors.RESET}  Connect to database
  {Colors.BRIGHT_BLUE}\\clear, \\cls{Colors.RESET}           Clear screen

{Colors.BOLD}Hot-Reload Commands:{Colors.RESET}
  {Colors.BRIGHT_GREEN}\\r, \\reload{Colors.RESET}           Reload all SQL engine modules
  {Colors.BRIGHT_GREEN}\\modules{Colors.RESET}               List loaded modules
  {Colors.BRIGHT_GREEN}\\debug{Colors.RESET}                 Toggle debug mode

{Colors.BOLD}Configuration:{Colors.RESET}
  {Colors.BRIGHT_YELLOW}\\timing{Colors.RESET}                Toggle query timing on/off
  {Colors.BRIGHT_YELLOW}\\echo{Colors.RESET}                  Toggle query echo on/off
  {Colors.BRIGHT_YELLOW}\\set <option> <value>{Colors.RESET}  Set configuration option
  {Colors.BRIGHT_YELLOW}\\show [option]{Colors.RESET}         Show configuration

{Colors.BOLD}Data Operations:{Colors.RESET}
  {Colors.BRIGHT_MAGENTA}\\export <format> <file>{Colors.RESET} Export last SELECT result
  {Colors.BRIGHT_MAGENTA}\\import <file>{Colors.RESET}         Import and execute SQL file
  {Colors.BRIGHT_MAGENTA}\\history{Colors.RESET}               Show command history

{Colors.BOLD}Information:{Colors.RESET}
  {Colors.BRIGHT_CYAN}\\version{Colors.RESET}                Show version information
  {Colors.BRIGHT_CYAN}\\status{Colors.RESET}                 Show session status

{Colors.BOLD}Configuration Options:{Colors.RESET}
  {Colors.WHITE}table_format{Colors.RESET}           ascii, markdown, csv
  {Colors.WHITE}max_column_width{Colors.RESET}       Maximum width for table columns
  {Colors.WHITE}null_display{Colors.RESET}           How to display NULL values
  {Colors.WHITE}colorize_output{Colors.RESET}        Enable/disable colors (true/false)
  {Colors.WHITE}auto_reload{Colors.RESET}            Auto-reload modules on file changes
  {Colors.WHITE}reload_on_error{Colors.RESET}        Auto-reload on execution errors

{Colors.BOLD}Tips:{Colors.RESET}
  • End queries with semicolon (;)
  • Multi-line queries supported
  • Ctrl+C cancels current query
  • Use ↑↓ arrows for history navigation
  • Tab completion for SQL keywords

{Colors.BOLD}Examples:{Colors.RESET}
  {Colors.DIM}SELECT * FROM users WHERE age > 25;{Colors.RESET}
  {Colors.DIM}\\set table_format markdown{Colors.RESET}
  {Colors.DIM}\\export csv /tmp/results.csv{Colors.RESET}
  {Colors.DIM}\\r  # Hot-reload after code changes{Colors.RESET}
"""
        print(help_text)
    
    def _cmd_list_databases(self, args):
        """List all databases with enhanced display"""
        print(f"{Colors.BOLD}Available Databases:{Colors.RESET}\n")
        
        try:
            if not hasattr(db_manager, 'databases') or not db_manager.databases:
                self._print_warning("No databases found")
                return
            
            databases_info = []
            for db_file in db_manager.databases:
                try:
                    filename = os.path.basename(db_file)
                    if filename.endswith('.su'):
                        filename = filename[:-3]
                    
                    # Determine status
                    if db_file == getattr(db_manager, 'active_db_name', None):
                        status = "Online"
                    else:
                        status = "Available"
                    
                    # Get file size
                    try:
                        size = os.path.getsize(db_file)
                        size_str = f"{size:,} bytes" if size < 1024 else f"{size//1024:,} KB"
                    except:
                        size_str = "Unknown"
                    
                    databases_info.append({
                        'Database': filename,
                        'Status': status,
                        'Size': size_str,
                        'Path': db_file
                    })
                
                except Exception as e:
                    databases_info.append({
                        'Database': str(db_file),
                        'Status': "ERROR",
                        'Size': "N/A",
                        'Path': str(e)
                    })
            
            if databases_info:
                formatted_table = self.formatter.format_table(databases_info, ['Database', 'Status', 'Size'])
                print(formatted_table)
            else:
                self._print_warning("No databases available")
        
        except Exception as e:
            self._print_error(f"Failed to list databases: {e}")
    
    def _cmd_describe_tables(self, args):
        """List tables in current database with enhanced info"""
        try:
            current_db = getattr(db_manager, 'active_db', None)
            if not current_db:
                self._print_warning("No database selected")
                print(f"Use {Colors.BOLD}\\c <database_name>{Colors.RESET} to connect to a database")
                return
            
            print(f"{Colors.BOLD}Tables in current database:{Colors.RESET}\n")
            
            tables_info = []
            for table_name, table_obj in current_db.items():
                try:
                    row_count = len(table_obj.rows) if hasattr(table_obj, 'rows') else 0
                    col_count = len(table_obj.schema) if hasattr(table_obj, 'schema') else 0
                    
                    # Get column info
                    columns = []
                    if hasattr(table_obj, 'schema'):
                        for col_name, col_type in table_obj.schema.items():
                            columns.append(f"{col_name}({col_type})")
                    
                    tables_info.append({
                        'Table': table_name,
                        'Rows': f"{row_count:,}",
                        'Columns': str(col_count),
                        'Schema': ', '.join(columns[:3]) + ('...' if len(columns) > 3 else '')
                    })
                
                except Exception as e:
                    tables_info.append({
                        'Table': table_name,
                        'Rows': 'Error',
                        'Columns': 'Error',
                        'Schema': str(e)[:50]
                    })
            
            if tables_info:
                formatted_table = self.formatter.format_table(tables_info, ['Table', 'Rows', 'Columns', 'Schema'])
                print(formatted_table)
            else:
                self._print_info("No tables found in current database")

        except Exception as e:
            self._print_error(f"Failed to describe tables: {e}")
    
    def _cmd_connect(self, args):
        """Connect to a database"""
        if not args:
            self._print_error("Usage: \\c <database_name>")
            return
        
        db_name = args[0]
        try:
            print(f"{Colors.BRIGHT_BLUE}Connecting to database '{db_name}'...{Colors.RESET}")
            self._execute_query(f"USE {db_name}")
        except Exception as e:
            self._print_error(f"Connection failed: {e}")
    
    def _cmd_reload(self, args):
        """Hot-reload all SQL engine modules"""
        print(f"{Colors.BRIGHT_YELLOW}Reloading SQL engine modules...{Colors.RESET}")
        
        try:
            reloaded, failed = reload_modules()
            
            if reloaded:
                self._print_success(f"Reloaded modules: {', '.join(reloaded)}")
                
                # Re-import updated objects
                try:
                    global db_manager, Lexer, Parser, execute
                    from engine import db_manager, Lexer, Parser
                    from executor import execute
                    self._print_info("Updated global references")
                except Exception as e:
                    self._print_warning(f"Could not update references: {e}")
            
            if failed:
                for module, error in failed:
                    self._print_error(f"Failed to reload {module}: {error}")
            
            if not reloaded and not failed:
                self._print_warning("No modules registered for reloading")
        
        except Exception as e:
            self._print_error(f"Reload failed: {e}")
            print(f"{Colors.BRIGHT_BLACK}{traceback.format_exc()}{Colors.RESET}")
    
    def _cmd_list_modules(self, args):
        """List loaded modules available for hot-reload"""
        print(f"{Colors.BOLD}Registered Modules:{Colors.RESET}\n")
        
        if not MODULE_REGISTRY:
            self._print_warning("No modules registered for hot-reload")
            return
        
        modules_info = []
        for name, module in MODULE_REGISTRY.items():
            try:
                module_file = getattr(module, '__file__', 'Unknown')
                module_time = datetime.fromtimestamp(os.path.getmtime(module_file)).strftime('%Y-%m-%d %H:%M:%S') if module_file != 'Unknown' else 'Unknown'
                
                modules_info.append({
                    'Module': name,
                    'File': os.path.basename(module_file) if module_file != 'Unknown' else 'Unknown',
                    'Modified': module_time
                })
            except Exception as e:
                modules_info.append({
                    'Module': name,
                    'File': 'Error',
                    'Modified': str(e)[:30]
                })
        
        formatted_table = self.formatter.format_table(modules_info, ['Module', 'File', 'Modified'])
        print(formatted_table)
    
    def _cmd_debug_mode(self, args):
        """Toggle debug mode"""
        # Toggle debug configuration
        debug_mode = not getattr(self.config, 'debug_mode', False)
        self.config.debug_mode = debug_mode
        
        if debug_mode:
            self._print_success("Debug mode enabled - detailed error reporting active")
            self.config.reload_on_error = True
        else:
            self._print_info("Debug mode disabled")
            self.config.reload_on_error = False
        
        self.config.save_config()
    
    def _cmd_toggle_timing(self, args):
        """Toggle query timing with visual feedback"""
        self.config.timing = not self.config.timing
        if self.config.timing:
            self._print_success("Query timing enabled")
        else:
            self._print_info("Query timing disabled")
        self.config.save_config()
    
    def _cmd_toggle_echo(self, args):
        """Toggle query echo"""
        self.config.echo_queries = not self.config.echo_queries
        if self.config.echo_queries:
            self._print_success("Query echo enabled")
        else:
            self._print_info("Query echo disabled")
        self.config.save_config()
    
    def _cmd_set_config(self, args):
        """Set configuration option with validation"""
        if len(args) < 2:
            self._print_error("Usage: \\set <option> <value>")
            print(f"{Colors.YELLOW}Available options: table_format, max_column_width, null_display, colorize_output, auto_reload, reload_on_error{Colors.RESET}")
            return
        
        option, value = args[0], args[1]
        
        try:
            if option == 'table_format' and value in ['ascii', 'markdown', 'csv']:
                self.config.table_format = value
                self._print_success(f"Table format set to: {value}")
            elif option == 'max_column_width':
                self.config.max_column_width = int(value)
                self._print_success(f"Max column width set to: {value}")
            elif option == 'null_display':
                self.config.null_display = value
                self._print_success(f"NULL display set to: '{value}'")
            elif option == 'colorize_output':
                self.config.colorize_output = value.lower() == 'true'
                self._print_success(f"Color output {'enabled' if self.config.colorize_output else 'disabled'}")
            elif option == 'auto_reload':
                self.config.auto_reload = value.lower() == 'true'
                self._print_success(f"Auto-reload {'enabled' if self.config.auto_reload else 'disabled'}")
            elif option == 'reload_on_error':
                self.config.reload_on_error = value.lower() == 'true'
                self._print_success(f"Reload on error {'enabled' if self.config.reload_on_error else 'disabled'}")
            else:
                self._print_error(f"Unknown option: {option}")
                return
            
            self.config.save_config()
        
        except ValueError as e:
            self._print_error(f"Invalid value: {e}")
    
    def _cmd_show_config(self, args):
        """Show configuration with enhanced display"""
        if args:
            option = args[0]
            if hasattr(self.config, option):
                value = getattr(self.config, option)
                print(f"{Colors.BRIGHT_CYAN}{option}{Colors.RESET} = {Colors.BRIGHT_WHITE}{value}{Colors.RESET}")
            else:
                self._print_error(f"Unknown option: {option}")
        else:
            print(f"{Colors.BOLD}⚙️  Current Configuration:{Colors.RESET}\n")
            
            config_items = [
                ('table_format', self.config.table_format),
                ('max_column_width', self.config.max_column_width),
                ('null_display', f"'{self.config.null_display}'"),
                ('colorize_output', self.config.colorize_output),
                ('timing', self.config.timing),
                ('echo_queries', self.config.echo_queries),
                ('auto_reload', getattr(self.config, 'auto_reload', True)),
                ('reload_on_error', getattr(self.config, 'reload_on_error', False)),
            ]
            
            config_data = [{'Option': opt, 'Value': str(val)} for opt, val in config_items]
            formatted_table = self.formatter.format_table(config_data, ['Option', 'Value'])
            print(formatted_table)
    
    def _cmd_show_history(self, args):
        """Show command history with enhanced display"""
        try:
            if PROMPT_TOOLKIT_AVAILABLE:
                # Get history from prompt_toolkit
                history_items = []
                for i, entry in enumerate(self.history.file_history.get_strings()):
                    history_items.append({'#': i+1, 'Command': entry[:80] + ('...' if len(entry) > 80 else '')})
                
                if history_items:
                    print(f"{Colors.BOLD}Command History:{Colors.RESET}\n")
                    # Show last 20 items
                    recent_items = history_items[-20:] if len(history_items) > 20 else history_items
                    formatted_table = self.formatter.format_table(recent_items, ['#', 'Command'])
                    print(formatted_table)
                    
                    if len(history_items) > 20:
                        print(f"\n{Colors.BRIGHT_BLACK}Showing last 20 of {len(history_items)} commands{Colors.RESET}")
                else:
                    self._print_info("No command history available")
            else:
                # Fallback to readline history
                try:
                    import readline
                    print(f"{Colors.BOLD}Command History:{Colors.RESET}\n")
                    for i in range(max(0, readline.get_current_history_length() - 20), readline.get_current_history_length()):
                        line = readline.get_history_item(i + 1)
                        if line:
                            print(f"{i+1:4}: {line}")
                except:
                    self._print_warning("History not available")
        except Exception as e:
            self._print_error(f"Could not retrieve history: {e}")
    
    def _cmd_clear(self, args):
        """Clear the screen"""
        os.system('cls' if os.name == 'nt' else 'clear')
        # Reprint a minimal banner
        print(f"{Colors.BRIGHT_GREEN}{Colors.RESET} - {self._get_connection_info()}\n")
    
    def _cmd_export(self, args):
        """Export last query result with enhanced options"""
        if len(args) < 2:
            self._print_error("Usage: \\export <format> <filename>")
            print(f"{Colors.YELLOW}Supported formats: csv, json, sql, xlsx{Colors.RESET}")
            return
        
        if not self.config.last_result:
            self._print_error("No query result to export. Run a SELECT query first.")
            return
        
        format_type = args[0].lower()
        filename = args[1]
        
        try:
            print(f"{Colors.BRIGHT_BLUE}Exporting {len(self.config.last_result)} rows to {filename}...{Colors.RESET}")
            
            if format_type == 'csv':
                self._export_csv(filename)
            elif format_type == 'json':
                self._export_json(filename)
            elif format_type == 'sql':
                self._export_sql(filename)
            elif format_type == 'xlsx':
                self._export_xlsx(filename)
            else:
                self._print_error(f"Unsupported format: {format_type}")
                print(f"{Colors.YELLOW}Supported formats: csv, json, sql, xlsx{Colors.RESET}")
                return
            
            row_count = len(self.config.last_result)
            file_size = os.path.getsize(filename) if os.path.exists(filename) else 0
            size_str = f"{file_size:,} bytes" if file_size < 1024 else f"{file_size//1024:,} KB"
            
            self._print_success(f"Exported {row_count:,} rows to {filename} ({format_type.upper()}, {size_str})")
            
        except Exception as e:
            self._print_error(f"Export failed: {str(e)}")
    
    def _export_csv(self, filename):
        """Export to CSV format"""
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            if not self.config.last_result:
                return
            
            columns = list(self.config.last_result[0].keys())
            writer = csv.writer(csvfile)
            
            writer.writerow(columns)
            
            for row in self.config.last_result:
                row_data = [row.get(col, '') for col in columns]
                writer.writerow(row_data)
    
    def _export_json(self, filename):
        """Export to JSON format"""
        if not filename.endswith('.json'):
            filename += '.json'
        
        with open(filename, 'w', encoding='utf-8') as jsonfile:
            json.dump({
                'exported_at': datetime.now().isoformat(),
                'row_count': len(self.config.last_result),
                'data': self.config.last_result
            }, jsonfile, indent=2, default=str)
    
    def _export_sql(self, filename):
        """Export as SQL INSERT statements"""
        if not filename.endswith('.sql'):
            filename += '.sql'
        
        if not self.config.last_result:
            return
        
        # Ask for table name with a default
        try:
            if PROMPT_TOOLKIT_AVAILABLE:
                table_name = prompt("Enter table name for INSERT statements (default: exported_data): ").strip()
            else:
                table_name = input("Enter table name for INSERT statements (default: exported_data): ").strip()
        except (KeyboardInterrupt, EOFError):
            table_name = ""
        
        if not table_name:
            table_name = "exported_data"
        
        with open(filename, 'w', encoding='utf-8') as sqlfile:
            columns = list(self.config.last_result[0].keys())
            
            # Write enhanced header
            sqlfile.write(f"-- MyShell SQL Export\n")
            sqlfile.write(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            sqlfile.write(f"-- Rows: {len(self.config.last_result):,}\n")
            sqlfile.write(f"-- Table: {table_name}\n\n")
            
            # Write INSERT statements
            for i, row in enumerate(self.config.last_result):
                values = []
                for col in columns:
                    value = row.get(col)
                    if value is None:
                        values.append('NULL')
                    elif isinstance(value, str):
                        escaped_value = value.replace("'", "''")
                        values.append(f"'{escaped_value}'")
                    elif isinstance(value, bool):
                        values.append('TRUE' if value else 'FALSE')
                    else:
                        values.append(str(value))
                
                columns_str = ', '.join(columns)
                values_str = ', '.join(values)
                sqlfile.write(f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});\n")
                
                # Add progress comments for large exports
                if i > 0 and i % 1000 == 0:
                    sqlfile.write(f"\n-- Progress: {i:,} rows inserted\n")
    
    def _export_xlsx(self, filename):
        """Export to Excel format (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self._print_error("Excel export requires openpyxl: pip install openpyxl")
            return
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyShell Export"
        
        if not self.config.last_result:
            return
        
        columns = list(self.config.last_result[0].keys())
        
        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, row in enumerate(self.config.last_result, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def _cmd_import(self, args):
        """Import SQL file with progress tracking"""
        if not args:
            self._print_error("Usage: \\import <filename>")
            return
        
        filename = args[0]
        
        if not os.path.exists(filename):
            self._print_error(f"File not found: {filename}")
            return
        
        try:
            print(f"{Colors.BRIGHT_BLUE}Importing SQL from {filename}...{Colors.RESET}")
            
            with open(filename, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Split into statements
            statements = [s.strip() for s in content.split(';') if s.strip()]
            
            if not statements:
                self._print_warning("No SQL statements found in file")
                return
            
            print(f"{Colors.BRIGHT_BLACK}Found {len(statements)} statements to execute...{Colors.RESET}")
            
            # Execute statements with progress
            successful = 0
            failed = 0
            
            for i, stmt in enumerate(statements, 1):
                try:
                    if len(stmt) > 50:
                        preview = stmt[:50] + "..."
                    else:
                        preview = stmt
                    
                    print(f"{Colors.DIM}[{i:3}/{len(statements)}] {preview}{Colors.RESET}")
                    self._execute_query(stmt + ';')
                    successful += 1
                    
                except Exception as e:
                    self._print_error(f"Statement {i} failed: {e}")
                    failed += 1
                    
                    # Ask user if they want to continue
                    if failed >= 3:
                        try:
                            if PROMPT_TOOLKIT_AVAILABLE:
                                continue_import = prompt(f"{Colors.YELLOW}Multiple failures detected. Continue? (y/n): {Colors.RESET}").lower()
                            else:
                                continue_import = input(f"{Colors.YELLOW}Multiple failures detected. Continue? (y/n): {Colors.RESET}").lower()
                            
                            if continue_import not in ['y', 'yes']:
                                break
                        except (KeyboardInterrupt, EOFError):
                            break
            
            # Summary
            total = successful + failed
            if successful > 0:
                self._print_success(f"Import completed: {successful}/{total} statements successful")
            if failed > 0:
                self._print_warning(f"{failed} statements failed")
                
        except Exception as e:
            self._print_error(f"Import failed: {e}")
    
    def _cmd_version(self, args):
        """Show comprehensive version information"""
        python_version = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        
        version_info = f"""

{Colors.BOLD}{Colors.RESET}
  Version: {Colors.BRIGHT_GREEN}2.1.0{Colors.RESET}
  
{Colors.BOLD}Runtime Environment{Colors.RESET}
  Python: {Colors.BRIGHT_BLUE}{python_version}{Colors.RESET}
  Platform: {Colors.BRIGHT_BLUE}{sys.platform}{Colors.RESET}
  
{Colors.BOLD}Features{Colors.RESET}
  Hot-reload: {Colors.BRIGHT_GREEN}Enabled{Colors.RESET}
  Enhanced UI: {Colors.BRIGHT_GREEN if PROMPT_TOOLKIT_AVAILABLE else Colors.BRIGHT_RED}{'Available' if PROMPT_TOOLKIT_AVAILABLE else 'Missing (pip install prompt-toolkit)'}{Colors.RESET}
  Export formats: {Colors.BRIGHT_BLUE}CSV, JSON, SQL, XLSX{Colors.RESET}
  
{Colors.BOLD}Modules Status{Colors.RESET}"""

        # Check module status
        module_status = []
        for name in ['database_manager', 'engine', 'executor', 'datatypes']:
            if name in MODULE_REGISTRY:
                module_status.append(f"  {name}: {Colors.BRIGHT_GREEN}Loaded{Colors.RESET}")
            else:
                module_status.append(f"  {name}: {Colors.BRIGHT_RED}Missing{Colors.RESET}")
        
        version_info += '\n' + '\n'.join(module_status)
        
        # Add session info
        uptime = datetime.now() - self.start_time
        version_info += f"""

{Colors.BOLD}Session Info{Colors.RESET}
  Uptime: {Colors.BRIGHT_BLUE}{uptime}{Colors.RESET}
  Queries executed: {Colors.BRIGHT_BLUE}{self.query_count:,}{Colors.RESET}
"""
        
        print(version_info)
    
    def _cmd_status(self, args):
        """Show comprehensive session and connection status"""
        uptime = datetime.now() - self.start_time
        
        # Get database info
        try:
            current_db = getattr(db_manager, 'active_db_name', None)
            if current_db:
                db_name = os.path.basename(current_db)
                if db_name.endswith('.su'):
                    db_name = db_name[:-3]
                db_status = f"{Colors.BRIGHT_GREEN}{db_name}{Colors.RESET}"
                
                # Get table count
                try:
                    table_count = len(db_manager.active_db)
                    db_info = f" ({table_count} tables)"
                except:
                    db_info = ""
            else:
                db_status = f"{Colors.BRIGHT_RED}Not connected{Colors.RESET}"
                db_info = ""
        except:
            db_status = f"{Colors.BRIGHT_RED}Database manager unavailable{Colors.RESET}"
            db_info = ""
        
        # Memory usage (if available)
        try:
            import psutil
            process = psutil.Process()
            memory_mb = process.memory_info().rss / 1024 / 1024
            memory_info = f"{memory_mb:.1f} MB"
        except:
            memory_info = "N/A"
        
        status_info = f"""

{Colors.BOLD}Database Connection{Colors.RESET}
  Current database: {db_status}{db_info}
  Available databases: {Colors.BRIGHT_BLUE}{len(getattr(db_manager, 'databases', []))}{Colors.RESET}

{Colors.BOLD}📈 Session Statistics{Colors.RESET}
  Queries executed: {Colors.BRIGHT_BLUE}{self.query_count:,}{Colors.RESET}
  Session uptime: {Colors.BRIGHT_BLUE}{uptime}{Colors.RESET}
  Memory usage: {Colors.BRIGHT_BLUE}{memory_info}{Colors.RESET}

{Colors.BOLD}⚙️  Configuration{Colors.RESET}
  Query timing: {Colors.BRIGHT_GREEN if self.config.timing else Colors.BRIGHT_RED}{'ON' if self.config.timing else 'OFF'}{Colors.RESET}
  Query echo: {Colors.BRIGHT_GREEN if self.config.echo_queries else Colors.BRIGHT_RED}{'ON' if self.config.echo_queries else 'OFF'}{Colors.RESET}
  Table format: {Colors.BRIGHT_BLUE}{self.config.table_format}{Colors.RESET}
  Auto-reload: {Colors.BRIGHT_GREEN if getattr(self.config, 'auto_reload', True) else Colors.BRIGHT_RED}{'ON' if getattr(self.config, 'auto_reload', True) else 'OFF'}{Colors.RESET}
  
{Colors.BOLD}Hot-Reload Status{Colors.RESET}
  Registered modules: {Colors.BRIGHT_BLUE}{len(MODULE_REGISTRY)}{Colors.RESET}
  Last result rows: {Colors.BRIGHT_BLUE}{len(self.config.last_result) if self.config.last_result else 0:,}{Colors.RESET}
"""
        print(status_info)


def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    """Enhanced main entry point with better argument handling"""
    parser = argparse.ArgumentParser(
        description='',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  myshell                          # Start interactive shell
  myshell --execute "SELECT * FROM users;"  # Execute query and exit
  myshell --no-color               # Disable colored output
  myshell --config /path/config.json       # Use custom config file
  
Hot-reload commands:
  \\r, \\reload                     # Reload SQL engine modules
  \\modules                        # List loaded modules
        """
    )
    
    parser.add_argument('--no-color', action='store_true', 
                       help='Disable colored output')
    parser.add_argument('--config', 
                       help='Path to config file')
    parser.add_argument('--execute', '-e', 
                       help='Execute command and exit')
    parser.add_argument('--version', action='version', version='MyShell SQL CLI 2.1.0')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug mode with detailed error reporting')
    parser.add_argument('--import-file', 
                       help='Import SQL file on startup')
    parser.add_argument('--export-config', 
                       help='Export current configuration to file')
    
    args = parser.parse_args()
    
    # Handle early exit operations
    if args.export_config:
        try:
            shell = EnhancedSQLShell()
            shell.config.save_config()
            print(f"Configuration exported to {shell.config.config_file}")
            return
        except Exception as e:
            print(f"Failed to export config: {e}")
            sys.exit(1)
    
    try:
        shell = EnhancedSQLShell()
        
        # Apply command line options
        if args.no_color:
            shell.config.colorize_output = False
        
        if args.debug:
            shell.config.debug_mode = True
            shell.config.reload_on_error = True
            print(f"{Colors.BRIGHT_YELLOW}🐛 Debug mode enabled{Colors.RESET}")
        
        if args.config:
            # Load custom config file
            try:
                shell.config.config_file = Path(args.config)
                shell.config.load_config()
                print(f"{Colors.BRIGHT_GREEN}Loaded config from {args.config}{Colors.RESET}")
            except Exception as e:
                print(f"{Colors.BRIGHT_RED}Failed to load config: {e}{Colors.RESET}")
        
        # Handle one-time operations
        if args.execute:
            shell._execute_query(args.execute)
            return
        
        if args.import_file:
            shell._cmd_import([args.import_file])
            if not sys.stdin.isatty():  # If running in script mode
                return
        
        # Start interactive shell
        shell.run()
        
    except KeyboardInterrupt:
        print(f"\n{Colors.BRIGHT_GREEN}Interrupted by user{Colors.RESET}")
        sys.exit(0)
    except Exception as e:
        print(f"{Colors.BRIGHT_RED}Fatal error: {e}{Colors.RESET}")
        if args.debug:
            print(f"{Colors.BRIGHT_BLACK}{traceback.format_exc()}{Colors.RESET}")
        sys.exit(1)


if __name__ == '__main__':
    main()
    
